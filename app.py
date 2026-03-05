import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import json
import os
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime, date, timedelta
import plotly.express as px
import plotly.graph_objects as go
from pathlib import Path
import time
from modules import automation as automated_reminders
from modules import db
import re
import io
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Initialize Database
db.init_db()

# === CONFIG ===================================================================
st.set_page_config(
    page_title="Sidekick CRM",
    page_icon="🎯",
    layout="wide",
    initial_sidebar_state="expanded"
)

# === CUSTOM CSS ===============================================================
def local_css(file_name):
    if os.path.exists(file_name):
        with open(file_name, encoding="utf8") as f:
            st.markdown(f'<style>{f.read()}</style>', unsafe_allow_html=True)

# local_css("assets/style.css") # Moved to after authentication check

import base64
def get_base64(bin_file):
    with open(bin_file, 'rb') as f:
        data = f.read()
    return base64.b64encode(data).decode()

def get_img_with_href(local_img_path, target_width="250px", extra_style=""):
    try:
        if os.path.exists(local_img_path):
            img_format = local_img_path.split(".")[-1]
            bin_str = get_base64(local_img_path)
            html_code = f'<img src="data:image/{img_format};base64,{bin_str}" style="max-width: {target_width}; height: auto; {extra_style}">'
            return html_code
    except:
        pass
    return None

# === SESSION STATE INITIALIZATION =============================================
if "authenticated" not in st.session_state:
    st.session_state.authenticated = False
if "username" not in st.session_state:
    st.session_state.username = None
if "role" not in st.session_state:
    st.session_state.role = "User"
if "allowed_pages" not in st.session_state:
    st.session_state.allowed_pages = "" 
if "show_loader" not in st.session_state:
    st.session_state.show_loader = False
if "logging_out" not in st.session_state:
    st.session_state.logging_out = False
if "pending_note" not in st.session_state:
    st.session_state.pending_note = None

# === LOGIN SYSTEM =============================================================
def login_page():
    local_css("assets/login.css")

    # Kill 'Press Enter to submit form' tooltip via components.html (actually runs JS)
    components.html("""
        <script>
        (function() {
            function removeHints() {
                const doc = window.parent.document;
                doc.querySelectorAll('small, p, div').forEach(el => {
                    if (el.textContent && el.textContent.trim() === 'Press Enter to submit form') {
                        el.style.cssText = 'display:none!important;visibility:hidden!important;height:0!important;overflow:hidden!important;';
                    }
                });
            }
            // Run immediately
            removeHints();
            // Watch for Streamlit re-injecting it
            const observer = new MutationObserver(() => removeHints());
            observer.observe(window.parent.document.body, {
                childList: true, subtree: true, characterData: true
            });
            // Also run on each focus event
            window.parent.document.addEventListener('focusin', () => removeHints(), true);
        })();
        </script>
    """, height=0)

    st.markdown("""
        <div class="sweet-circle sc-1"></div>
        <div class="sweet-circle sc-2"></div>
        <div class="sweet-circle sc-3"></div>
        <div class="sweet-circle sc-dot1"></div>
        <div class="sweet-circle sc-dot2"></div>
        <style>
            .sweet-circle { position:fixed; border-radius:50%; pointer-events:none; z-index:3; }
            .sc-1 { top:35%; right:-50px; width:180px; height:180px; background:linear-gradient(135deg, #1b6656, #2d8a76); opacity:0.4; animation: float 14s infinite alternate ease-in-out; }
            .sc-2 { top:28%; right:-30px; width:140px; height:140px; background:linear-gradient(135deg, #1d4354, #2c637a); opacity:0.5; animation: float 12s infinite alternate-reverse ease-in-out; }
            .sc-3 { top:48%; right:100px; width:80px; height:80px; background:#7bb06b; opacity:0.6; animation: float 9s infinite alternate ease-in-out 1s; }
            .sc-dot1 { bottom:120px; right:120px; width:35px; height:35px; background:#7bb06b; opacity:0.6; animation: float 7s infinite alternate ease-in-out; }
            .sc-dot2 { top:150px; left:80px; width:25px; height:25px; background:#1b6656; opacity:0.3; animation: float 10s infinite alternate-reverse ease-in-out; }
            @keyframes float {
                0% { transform: translate(0, 0) scale(1); }
                100% { transform: translate(15px, -15px) scale(1.05); }
            }
        </style>
    """, unsafe_allow_html=True)

    _, center_col, _ = st.columns([1, 1.4, 1])
    with center_col:
        with st.form("login_form"):
            # Wide Sidekick Logo
            st.markdown(f"""
                <div class="sidekick-logo-container">
                    {get_img_with_href("assets/SDK_LOGO.png", "260px") or
                     '<div style="font-family:Inter; font-size:2rem; font-weight:900; color:#1b6656; letter-spacing:-0.03em;">⬟ Sidekick CRM</div>'}
                </div>

                <div class="login-title-main">Login</div>
                <div class="login-title-underline"></div>

                <div class="login-subtitle">
                    Welcome back! Access your CRM dashboard.
                </div>
            """, unsafe_allow_html=True)

            # Minimalist Underline Inputs
            user = st.text_input("Username", placeholder="Enter your username")
            pw   = st.text_input("Password", type="password", placeholder="Enter your password")

            # JavaScript Enter-to-Next focus logic
            components.html("""
                <script>
                (function() {
                    const doc = window.parent.document;
                    const check = setInterval(() => {
                        const inputs = Array.from(doc.querySelectorAll('input'));
                        const u = inputs.find(i => i.placeholder === 'Enter your username');
                        const p = inputs.find(i => i.placeholder === 'Enter your password');
                        if (u && p) {
                            clearInterval(check);
                            u.addEventListener('keydown', e => {
                                if (e.key === 'Enter') { e.preventDefault(); p.focus(); }
                            });
                        }
                    }, 500);
                })();
                </script>
            """, height=0)

            # Gradient Pill Button
            if st.form_submit_button("→  CONTINUE", use_container_width=True):
                user_record = db.verify_user(user, pw)
                if user_record:
                    st.session_state.authenticated    = True
                    st.session_state.username         = user_record['username']
                    st.session_state.role             = user_record['role']
                    st.session_state.allowed_pages    = user_record['allowed_pages']
                    st.session_state.show_loader      = True
                    st.rerun()
                else:
                    st.error("Authentication Failed: Invalid Credentials")

            st.markdown("""
                <div class="login-footer">
                    © 2026 Sidekick CRM. All rights reserved.
                </div>
            """, unsafe_allow_html=True)

if not st.session_state.authenticated:
    login_page()
    st.stop()

# Load main app styles AFTER login
if not st.session_state.logging_out:
    local_css("assets/style.css")

# Loading animations removed as per user request
def display_premium_loader():
    local_css("assets/premium_loader.css")
    st.markdown("""
        <div class="premium-loader-wrapper">
            <div class="loader-container">
                <div class="loader-ring"></div>
                <div class="loader-ring"></div>
                <div class="loader-ring"></div>
                <div class="loader-center"></div>
            </div>
            <div class="loader-text">Initialize Intelligence</div>
        </div>
    """, unsafe_allow_html=True)

def display_custom_notification(message, title="Protocol Update", note_type="success"):
    local_css("assets/notifications.css")
    icon = "✅" if note_type == "success" else ("❌" if note_type == "error" else "⚠️")
    st.markdown(f"""
        <div class="notification-container">
            <div class="notification-box {note_type}">
                <div class="notification-icon">{icon}</div>
                <div class="notification-content">
                    <div class="notification-title">{title}</div>
                    <div class="notification-message">{message}</div>
                </div>
                <div class="notification-progress">
                    <div class="notification-progress-bar"></div>
                </div>
            </div>
        </div>
    """, unsafe_allow_html=True)

# Show pending notifications
if st.session_state.get("pending_note"):
    n = st.session_state.pending_note
    display_custom_notification(n['msg'], n['title'], n['type'])
    st.session_state.pending_note = None

# Handle Logout (Instant)
if st.session_state.logging_out:
    st.session_state.authenticated = False
    st.session_state.logging_out = False
    st.rerun()

# Data Sync Handler (With Premium Loader)
if any(x not in st.session_state for x in ["leads", "tasks", "sales", "recurring_clients"]) or st.session_state.show_loader:
    display_premium_loader()
    
    # Actually load data from DB
    if st.session_state.show_loader:
        # Reduced delay or no delay for instant feel
        pass 

    st.session_state.leads = db.get_all_leads()
    st.session_state.tasks = db.get_all_tasks()
    st.session_state.sales = db.get_all_sales()
    st.session_state.settings = db.get_settings({
        "smtp_host": "smtp.gmail.com",
        "smtp_port": 587,
        "smtp_user": "",
        "smtp_pass": "",
        "notify_email": "",
        "gemini_api_key": "",
        "auto_reminders": True,
        "last_auto_run": ""
    })
    st.session_state.recurring_clients = db.get_recurring_clients()

    if st.session_state.show_loader:
        st.session_state.show_loader = False
        st.rerun()

# === AUTOMATED TASKS ==========================================================
def run_daily_checks():
    s = st.session_state.settings
    if not s.get("auto_reminders", False):
        return
        
    today = str(date.today())
    if s.get("last_auto_run") != today:
        # Avoid running multiple times if someone reloads fast
        st.session_state.settings["last_auto_run"] = today
        db.save_settings(st.session_state.settings)
        
        # Run in background if possible, or just run it (it's fast enough)
        try:
            # We call the function from the module
            automated_reminders.run_reminders()
        except Exception as e:
            print(f"Error in automated reminders: {e}")

run_daily_checks()

# --- OPTIMIZED DATA PROCESSING ENGINE ---
@st.cache_data
def get_processed_data(raw_leads, raw_tasks, selected_year, selected_month):
    """Extreme performance data engine."""
    df_l = pd.DataFrame(raw_leads)
    df_t = pd.DataFrame(raw_tasks)
    
    lead_map = {l['id']: l['name'] for l in raw_leads}
    
    if df_l.empty:
        available_years = [datetime.now().year]
        filtered_leads = []
    else:
        df_l['dt'] = pd.to_datetime(df_l['created_at'], errors='coerce')
        df_l['year'] = df_l['dt'].dt.year.fillna(0).astype(int)
        df_l['month'] = df_l['dt'].dt.month.fillna(0).astype(int)
        available_years = sorted([y for y in df_l['year'].unique() if y > 0], reverse=True) or [datetime.now().year]
        
        f_df_l = df_l[df_l['year'] == selected_year]
        if selected_month > 0:
            f_df_l = f_df_l[f_df_l['month'] == selected_month]
        filtered_leads = f_df_l.to_dict('records')

    if df_t.empty:
        filtered_tasks = []
    else:
        df_t['dt'] = pd.to_datetime(df_t['created_at'], errors='coerce')
        df_t['year'] = df_t['dt'].dt.year.fillna(0).astype(int)
        df_t['month'] = df_t['dt'].dt.month.fillna(0).astype(int)
        
        f_df_t = df_t[df_t['year'] == selected_year]
        if selected_month > 0:
            f_df_t = f_df_t[f_df_t['month'] == selected_month]
        filtered_tasks = f_df_t.to_dict('records')

    # Pre-calculate Metrics (O(1) approach where possible)
    total_l = len(filtered_leads)
    open_t = sum(1 for t in filtered_tasks if not t.get("done"))
    
    today = date.today()
    overdue_count = 0
    for t in filtered_tasks:
        if not t.get("done") and t.get("due_date"):
            try:
                if datetime.strptime(t["due_date"], "%Y-%m-%d").date() < today:
                    overdue_count += 1
            except: pass
            
    return {
        "available_years": available_years,
        "leads": filtered_leads,
        "tasks": filtered_tasks,
        "df_leads": f_df_l,
        "df_tasks": f_df_t,
        "lead_map": lead_map,
        "stats": {
            "total_leads": total_l,
            "open_tasks": open_t,
            "overdue": overdue_count
        }
    }

# --- HELPERS ---
def get_lead_name(lead_id, lead_map=None):
    if lead_map:
        return lead_map.get(lead_id, "Unknown")
    # Fallback if map not provided
    for l in st.session_state.leads:
        if l["id"] == lead_id:
            return l["name"]
    return "Unknown"

STATUS_COLORS = {"New": "badge-new", "In Progress": "badge-inprogress", "Closed": "badge-closed"}
TEMP_COLORS = {"Hot": "badge-hot", "Warm": "badge-warm", "Cold": "badge-cold"}

def get_premium_email_layout(subject, body):
    styled_body = body.replace("\n", "<br>") 
    html = f"""
    <!-- email-wrapped -->
    <!DOCTYPE html>
    <html>
    <head>
    <style>
        .email-wrapper {{ background-color: #f1f5f9; padding: 40px 10px; font-family: sans-serif; }}
        .email-card {{ max-width: 600px; margin: 0 auto; background: #ffffff; border-radius: 20px; overflow: hidden; box-shadow: 0 10px 40px rgba(0,0,0,0.08); }}
        .header {{ background: linear-gradient(135deg, #1b6656 0%, #1d4354 100%); padding: 40px 20px; text-align: center; color: white; }}
        .content {{ padding: 45px; line-height: 1.7; font-size: 16px; color: #1d4354; }}
        .footer {{ padding: 25px; background: #f8fafc; text-align: center; font-size: 11px; color: #64748b; }}
    </style>
    </head>
    <body>
    <div class="email-wrapper"><div class="email-card"><div class="header"><h1>{subject}</h1></div>
    <div class="content">{styled_body}</div><div class="footer">&copy; {datetime.now().year} Sidekick Intelligence Protocol</div></div></div>
    </body></html>
    """
    return html

def send_email(subject, body, to_email=None):
    s = st.session_state.settings
    to = to_email or s.get("notify_email", "")
    if not (body.strip().startswith("<") or "<html>" in body.lower()):
        body = get_premium_email_layout(subject, body)
    if not all([s.get("smtp_user"), s.get("smtp_pass"), to]):
        return False, "SMTP not configured"
    try:
        msg = MIMEMultipart()
        msg["From"] = s["smtp_user"]; msg["To"] = to; msg["Subject"] = subject
        msg.attach(MIMEText(body, "html"))
        with smtplib.SMTP(s["smtp_host"], int(s["smtp_port"])) as server:
            server.starttls(); server.login(s["smtp_user"], s["smtp_pass"])
            server.send_message(msg)
        return True, "Email sent!"
    except Exception as e: return False, str(e)

# ─── NAVIGATION & GLOBAL FILTERS ───
# ─── NAVIGATION & GLOBAL FILTERS ───
if not st.session_state.logging_out:
    with st.sidebar:
        st.markdown(f"""
        <div style='padding: 15px 5px 25px 5px; margin-top: -45px; text-align:center;'>
            <div style='margin-bottom: 8px;'>
                {get_img_with_href("assets/SDK_LOGO.png", "225px", "filter: brightness(0) invert(1) drop-shadow(0 0 18px rgba(27, 102, 86, 0.4));") or '<h2 style="color:white; font-family:Outfit; font-size:2.2rem; font-weight:900; letter-spacing:-0.05em; margin:0;">SIDEKICK</h2>'}
            </div>
            <div style="font-size: 0.58rem; color: #7bb06b; font-weight: 800; letter-spacing: 0.3em; text-transform: uppercase; opacity: 1; margin-top: 5px;">Enterprise CRM</div>
        </div>
        """, unsafe_allow_html=True)

        # Initial year check to avoid bootstrap error
        years_temp = [datetime.now().year]
        
        st.markdown('<div class="filter-container" style="background:rgba(255,255,255,0.05); padding:12px; border-radius:12px; margin-bottom:15px; border:1px solid rgba(255,255,255,0.1);">', unsafe_allow_html=True)
        temp_sel_year = st.selectbox("Year Filter", options=[datetime.now().year], key="temp_y", label_visibility="collapsed") # Placeholder
        month_names = ["Full Year Intel", "January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
        selected_month_name = st.selectbox("Month Filter", options=month_names, index=0)
        selected_month = month_names.index(selected_month_name)
        st.markdown('</div>', unsafe_allow_html=True)

        # PROCESS DATA (CACHED)
        raw_l = st.session_state.leads
        raw_t = st.session_state.tasks
        
        # We need a first pass to get years
        data_bundle = get_processed_data(raw_l, raw_t, datetime.now().year, 0)
        available_years = data_bundle["available_years"]
        
        # Correct the selectbox with real years
        st.markdown("""<style>div[data-testid="stSelectbox"]+div { margin-top: -85px !important; }</style>""", unsafe_allow_html=True) # visual fix
        # RE-DOING Filtering logic cleaner
        selected_year = st.selectbox("Select Year", options=available_years, index=0, key="nav_year")
        
        # Get final data
        data = get_processed_data(raw_l, raw_t, selected_year, selected_month)
        leads = data["leads"]
        tasks = data["tasks"]
        df_l_active = data["df_leads"]
        df_t_active = data["df_tasks"]
        stats = data["stats"]
        l_map = data["lead_map"]

        all_nav_items = ["📊 Dashboard", "💰 Sales Report", "👥 Leads", "✅ Tasks", "📧 Reminders", "📢 Email Marketing", "⚙️ Settings", "👤 User Management"]
        if st.session_state.role == "Admin": nav_options = all_nav_items
        else:
            allowed = st.session_state.allowed_pages.split(",") if st.session_state.allowed_pages else []
            nav_options = [item for item in all_nav_items if item in allowed] or ["🚫 Access Restricted"]

        page = st.radio("Navigate", nav_options, label_visibility="collapsed")
        
        st.markdown("<hr style='margin: 10px 0;'>", unsafe_allow_html=True)
        st.markdown(f"""
        <div style='padding: 10px 15px;'>
        <div class="luxury-container" style="margin-bottom: 12px; border-color: rgba(123, 176, 107, 0.3); background: rgba(255,255,255,0.05); padding: 15px;">
        <div style="font-size: 0.7rem; color: #ffffff; text-transform: uppercase; letter-spacing: 0.1em; margin-bottom: 5px; opacity: 0.9;">Active Leads</div>
        <div style="font-size: 1.6rem; font-weight: 800; color: #ffffff;">{stats['total_leads']}</div>
        </div>
        <div class="luxury-container" style="background: rgba(255,255,255,0.03); border-color: rgba(123, 176, 107, 0.15); padding: 15px;">
        <div style="font-size: 0.7rem; color: #ffffff; text-transform: uppercase; letter-spacing: 0.1em; margin-bottom: 5px; opacity: 0.9;">Open Tasks</div>
        <div style="font-size: 1.6rem; font-weight: 800; color: #ffffff;">{stats['open_tasks']}</div>
        {f'<div style="font-size: 0.65rem; color: #ff9999; margin-top: 5px; font-weight: 600;">⚠️ {stats["overdue"]} OVERDUE</div>' if stats["overdue"] > 0 else ''}
        </div>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("<div style='height: 20px;'></div>", unsafe_allow_html=True)
        if st.button("🚪 LOGOUT ACCESS", key="logout_btn", use_container_width=True):
            st.session_state.logging_out = True
            st.rerun()
else:
    # Set default variables so the dashboard logic doesn't crash during logout animation
    leads = st.session_state.get('leads', [])
    tasks = st.session_state.get('tasks', [])
    stats = {'total_leads': 0, 'open_tasks': 0, 'overdue': 0}
    page = ""
    selected_year = datetime.now().year
    selected_month_name = "Full Year Intel"
    selected_month = 0
    df_l_active = pd.DataFrame()
    df_t_active = pd.DataFrame()


def confirm_delete_user(usr_id, username):
    st.warning(f"Are you sure you want to delete user **{username}**?")
    if st.button(f"Yes, Delete {username}", key=f"del_u_{usr_id}"):
        db.delete_user(usr_id)
        st.toast(f"User {username} deleted.")
        st.rerun()

def confirm_delete_lead(lead_id, name):
    st.warning(f"Are you sure you want to delete lead **{name}**?")
    if st.button(f"Yes, Delete {name}", key=f"del_l_{lead_id}"):
        db.delete_lead(lead_id)
        st.session_state.leads = db.get_all_leads()
        st.toast(f"Lead {name} deleted.")
        st.rerun()

# Add/Edit Dialogs removed as per revert request


# ═══════════════════════════════════════════════════════════════════════════════
# DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════════
if page == "📊 Dashboard":
    st.markdown("""
    <div class='page-header' style='margin-bottom: 40px;'>
        <div style='display:flex; justify-content:space-between; align-items:center;'>
            <div>
                <div class='page-title'>Command Center</div>
                <div class='page-sub'>Unified Lead Intelligence & Task Tactical Operations</div>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown(f"""
        <div style="margin-bottom:20px; font-size:0.85rem; color:#1b6656; font-weight:700; opacity:0.8;">
            📊 Current Analysis Period: <span style="color:#1d4354">{selected_year}</span> {f' - <span style="color:#1d4354">{selected_month_name}</span>' if selected_month > 0 else ''}
        </div>
    """, unsafe_allow_html=True)
    
    # Premium KPI Stats
    active = sum(1 for l in leads if l.get("status") == "New" or l.get("status") == "In Progress")
    pending_tasks = sum(1 for t in tasks if not t.get("done"))
    conversion = (sum(1 for l in leads if l.get("status") == "Closed") / len(leads) * 100) if leads else 0
    cold_leads = sum(1 for l in leads if l.get("temperature") == "Cold")
    hot_leads = sum(1 for l in leads if l.get("temperature") == "Hot")

    st.markdown(f"""
    <div class="kpi-container">
        <div class="kpi-card" style="--accent-gradient: linear-gradient(90deg, #1b6656, #7bb06b);">
            <div class="kpi-icon">👥</div>
            <div class="kpi-value" style="color:#1b6656">{len(leads)}</div>
            <div class="kpi-label">Total Portfolio</div>
        </div>
        <div class="kpi-card" style="--accent-gradient: linear-gradient(90deg, #1b6656, #1d4354);">
            <div class="kpi-icon">⚡</div>
            <div class="kpi-value" style="color:#1b6656">{active}</div>
            <div class="kpi-label">Active Pipeline</div>
        </div>
        <div class="kpi-card" style="--accent-gradient: linear-gradient(90deg, #7bb06b, #1b6656);">
            <div class="kpi-icon">🎯</div>
            <div class="kpi-value" style="color:#1d4354">{conversion:.1f}%</div>
            <div class="kpi-label">Closure Velocity</div>
        </div>
        <div class="kpi-card" style="--accent-gradient: linear-gradient(90deg, #fbbf24, #ef4444);">
            <div class="kpi-icon">🔥</div>
            <div class="kpi-value" style="color:#ef4444">{hot_leads}</div>
            <div class="kpi-label">High-Value (Hot)</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # --- ADVANCED ANALYTICS SECTION ---
    st.markdown('<div class="section-heading">🧠 Global Intelligence Pulse</div>', unsafe_allow_html=True)
    
    row1_c1, row1_c2 = st.columns([3, 2])
    
    @st.cache_resource
    def get_funnel_chart(df):
        stage_order = ["New", "In Progress", "Closed"]
        pipe_counts = df["status"].value_counts().reindex(stage_order).fillna(0).reset_index()
        pipe_counts.columns = ["Stage", "Count"]
        fig = go.Figure(go.Funnel(
            y=pipe_counts["Stage"], x=pipe_counts["Count"],
            textposition = "inside", textinfo = "value+percent initial",
            marker = {"color": ["#1b6656", "#1d4354", "#7bb06b"], "line": {"width": [4, 2, 2], "color": ["rgba(255,255,255,0.2)"] * 3}},
            connector = {"line": {"color": "rgba(255,255,255,0.1)", "width": 3}}
        ))
        fig.update_layout(template="plotly_white", paper_bgcolor="white", plot_bgcolor="white",
            font=dict(family="Outfit", color="#1d4354", size=12), margin=dict(l=20, r=20, t=60, b=20), height=400,
            title=dict(text="EXECUTIVE PIPELINE VELOCITY", x=0.5, y=0.95, font=dict(size=16, weight=800, color="#1b6656")))
        return fig

    @st.cache_resource
    def get_source_chart(df):
        source_counts = df["source"].value_counts().reset_index()
        source_counts.columns = ["Source", "Count"]
        fig = px.treemap(source_counts, path=["Source"], values="Count", color="Count", color_continuous_scale='Mint')
        fig.update_layout(template="plotly_white", paper_bgcolor="white", plot_bgcolor="white",
            margin=dict(t=60, l=10, r=10, b=10), font=dict(family="Outfit", color="#1d4354"), height=400,
            title=dict(text="SOURCE ORIGIN INTEL", x=0.5, y=0.95, font=dict(size=16, weight=800, color="#1b6656")))
        fig.update_coloraxes(showscale=False)
        return fig

    with row1_c1:
        if not df_l_active.empty:
            st.markdown('<div class="glass-card" style="padding:10px !important;">', unsafe_allow_html=True)
            st.plotly_chart(get_funnel_chart(df_l_active), use_container_width=True, config={'displayModeBar': False})
            st.markdown('<div style="padding:0 15px 10px; font-size:0.75rem; color:#475569; opacity:0.8; line-height:1.4;"><b>Intel Note:</b> Pipeline velocity track.</div>', unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)
        else: st.info("System data pending...")

    with row1_c2:
        if not df_l_active.empty:
            st.markdown('<div class="glass-card" style="padding:10px !important;">', unsafe_allow_html=True)
            st.plotly_chart(get_source_chart(df_l_active), use_container_width=True, config={'displayModeBar': False})
            st.markdown('<div style="padding:0 15px 10px; font-size:0.75rem; color:#475569; opacity:0.8; line-height:1.4;"><b>Intel Note:</b> Source analysis.</div>', unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)

    # Row 2: Lead Pulse & Engagement
    row2_c1, row2_c2 = st.columns([2, 3])
    
    @st.cache_resource
    def get_temp_chart(df):
        temp_counts = df["temperature"].value_counts().reset_index()
        temp_counts.columns = ["Temp", "Count"]
        fig = px.pie(temp_counts, values="Count", names="Temp", hole=0.7, color="Temp", color_discrete_map={"Hot": "#ef4444", "Warm": "#f59e0b", "Cold": "#1b6656"})
        fig.update_layout(template="plotly_white", paper_bgcolor="white", showlegend=True,
            legend=dict(orientation="v", yanchor="middle", y=0.5, xanchor="left", x=1.05, font=dict(size=10)),
            margin=dict(t=80, l=50, r=120, b=80), height=350,
            title=dict(text="CLIMATE ANALYSIS", x=0.5, y=0.95, font=dict(size=16, weight=800, color="#1b6656")),
            annotations=[dict(text="TEMP", x=0.5, y=0.5, font_size=18, showarrow=False, font_family="Outfit", font_color="#1d4354")])
        fig.update_traces(textposition='outside')
        return fig

    @st.cache_resource
    def get_pulse_chart(df):
        pulse_data = df.groupby('dt').size().reset_index(name='Count')
        pulse_data.columns = ['Time Cycle', 'Ingestion Rate']
        fig = px.area(pulse_data, x='Time Cycle', y='Ingestion Rate')
        fig.update_traces(line_color='#1b6656', fillcolor='rgba(27, 102, 86, 0.1)')
        fig.update_layout(template="plotly_white", paper_bgcolor="white", plot_bgcolor="white", font=dict(family="Outfit", color="#1d4354"),
            xaxis=dict(showgrid=False, title=None), yaxis=dict(showgrid=True, gridcolor='rgba(27, 102, 86, 0.05)', title=None),
            margin=dict(t=60, l=40, r=20, b=40), height=350,
            title=dict(text="ENGAGEMENT DYNAMICS", x=0.5, y=0.95, font=dict(size=16, weight=800, color="#1b6656")))
        return fig

    with row2_c1:
        if not df_l_active.empty:
            st.markdown('<div class="glass-card" style="padding:10px !important;">', unsafe_allow_html=True)
            st.plotly_chart(get_temp_chart(df_l_active), use_container_width=True, config={'displayModeBar': False})
            st.markdown('</div>', unsafe_allow_html=True)

    with row2_c2:
        if not df_l_active.empty:
            st.markdown('<div class="glass-card" style="padding:10px !important;">', unsafe_allow_html=True)
            st.plotly_chart(get_pulse_chart(df_l_active), use_container_width=True, config={'displayModeBar': False})
            st.markdown("""
                <div style="padding:0 15px 10px; font-size:0.75rem; color:#475569; opacity:0.8; line-height:1.4;">
                    <b>Intel Note:</b> System ki "dharhkan" (pulse). Ye last 30 days mein nayi leads ki entry ka pattern dikhata hai.
                </div>
            """, unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)

    # --- TASK SPECIFIC INTELLIGENCE ---
    st.markdown('<div class="section-heading">🛠️ Task Tactical Load</div>', unsafe_allow_html=True)
    t_col1, t_col2 = st.columns([1, 1])
    
    @st.cache_resource
    def get_task_prio_chart(df):
        prio_counts = df["priority"].value_counts().reset_index()
        prio_counts.columns = ["Priority", "Count"]
        fig = px.pie(prio_counts, values="Count", names="Priority", hole=0.7, color="Priority", color_discrete_map={"High": "#ef4444", "Medium": "#fbbf24", "Low": "#10b981"})
        fig.update_layout(template="plotly_white", paper_bgcolor="white", showlegend=True,
            legend=dict(orientation="v", yanchor="middle", y=0.5, xanchor="left", x=1.05, font=dict(size=10)),
            margin=dict(t=80, l=40, r=120, b=80), height=350,
            title=dict(text="TASK OPERATIONAL LOAD", x=0.5, y=0.95, font=dict(size=14, weight=800, color="#1b6656")),
            annotations=[dict(text="PRIO", x=0.5, y=0.5, font_size=16, showarrow=False, font_family="Outfit", font_color="#1d4354")])
        fig.update_traces(textposition='outside')
        return fig

    with t_col1:
        if not df_t_active.empty:
            st.markdown('<div class="glass-card" style="padding:10px !important;">', unsafe_allow_html=True)
            st.plotly_chart(get_task_prio_chart(df_t_active), use_container_width=True, config={'displayModeBar': False})
            st.markdown('<div style="padding:0 15px 10px; font-size:0.75rem; color:#475569; opacity:0.8; line-height:1.4;"><b>Intel Note:</b> Task priority analysis.</div>', unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)

    with t_col2:
        if tasks:
            done_tasks = sum(1 for t in tasks if t.get("done"))
            total_t = len(tasks)
            
            fig_progress = go.Figure(go.Indicator(
                mode = "gauge+number",
                value = (done_tasks/total_t*100) if total_t > 0 else 0,
                domain = {'x': [0, 1], 'y': [0, 1]},
                title = {'text': "EXECUTION EFFICIENCY", 'font': {'size': 14, 'color': '#1b6656', 'family': 'Outfit'}},
                gauge = {
                    'axis': {'range': [None, 100], 'tickwidth': 1, 'tickcolor': "#1b6656"},
                    'bar': {'color': "#1b6656"},
                    'bgcolor': "white",
                    'borderwidth': 2,
                    'bordercolor': "rgba(27,102,86,0.1)",
                    'steps': [
                        {'range': [0, 50], 'color': 'rgba(239, 68, 68, 0.1)'},
                        {'range': [50, 80], 'color': 'rgba(251, 191, 36, 0.1)'},
                        {'range': [80, 100], 'color': 'rgba(16, 185, 129, 0.1)'}],
                }
            ))
            fig_progress.update_layout(paper_bgcolor="white", height=300, margin=dict(t=80, b=20))
            
            st.markdown('<div class="glass-card" style="padding:10px !important;">', unsafe_allow_html=True)
            st.plotly_chart(fig_progress, use_container_width=True, config={'displayModeBar': False})
            st.markdown("""
                <div style="padding:0 15px 10px; font-size:0.75rem; color:#475569; opacity:0.8; line-height:1.4;">
                    <b>Intel Note:</b> Task efficiency meter. Jitna ye 100% ke qareeb hoga, utna hi aapka kaam waqt par khatam ho raha hai.
                </div>
            """, unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="section-heading">🕒 Recent Leads Intelligence</div>', unsafe_allow_html=True)
    if leads:
        sorted_leads = sorted(leads, key=lambda x: x.get("created_at", ""), reverse=True)
        # Dashboard Pagination for Recent Leads
        DLR_PS = 5
        total_dlr = (len(sorted_leads) // DLR_PS) + (1 if len(sorted_leads) % DLR_PS > 0 else 0)
        if 'dash_leads_page' not in st.session_state: st.session_state.dash_leads_page = 1
        
        if total_dlr > 1:
            dc1, dc2, dc3, dc4 = st.columns([1,1,1,5])
            with dc1:
                if st.button("⬅️", key="dlr_prev", disabled=st.session_state.dash_leads_page <= 1):
                    st.session_state.dash_leads_page -= 1
                    st.rerun()
            with dc2:
                st.markdown(f"<div style='padding-top:10px; font-weight:700; color:#1b6656; font-size:0.8rem;'>{st.session_state.dash_leads_page}/{total_dlr}</div>", unsafe_allow_html=True)
            with dc3:
                if st.button("➡️", key="dlr_next", disabled=st.session_state.dash_leads_page >= total_dlr):
                    st.session_state.dash_leads_page += 1
                    st.rerun()
            
            s_idx = (st.session_state.dash_leads_page - 1) * DLR_PS
            recent = sorted_leads[s_idx:s_idx + DLR_PS]
        else:
            recent = sorted_leads[:5]

        for l in recent:
            sc = STATUS_COLORS.get(l.get("status", "New"), "badge-new")
            tc = TEMP_COLORS.get(l.get("temperature", "Warm"), "badge-warm")
            st.markdown(f"""
            <div class="glass-card" style="padding: 18px 24px; margin-bottom: 12px; border-left: 4px solid var(--primary);">
                <div style="display:flex; justify-content:space-between; align-items:flex-start;">
                    <div>
                        <div style="font-family:'Outfit',sans-serif; font-size:1.1rem; font-weight:700; color:#1d4354; margin-bottom:4px;">{l['name']}</div>
                        <div style="color:#1b6656; font-size:0.82rem; font-weight:500;">
                            <span style="display:inline-flex; align-items:center; gap:5px;">🏢 {l.get('company','Private Enterprise')}</span>
                        </div>
                    </div>
                    <div style="display:flex; gap:10px;">
                        <span class="badge {sc}">{l.get('status','New')}</span>
                        <span class="badge {tc}">{l.get('temperature','Warm')}</span>
                    </div>
                </div>
                <div style="height:1px; background:rgba(27, 102, 86, 0.1); margin:12px 0;"></div>
                <div style="display:flex; gap:20px; color:#475569; font-size:0.78rem; opacity:0.8;">
                    <span style="display:inline-flex; align-items:center; gap:6px;">📧 {l.get('email','-')}</span>
                    <span style="display:inline-flex; align-items:center; gap:6px;">📞 {l.get('phone','-')}</span>
                    <span style="margin-left:auto; color:rgba(27, 102, 86, 0.45);">Captured: {l.get('created_at','')[:10]}</span>
                </div>
            </div>
            """, unsafe_allow_html=True)
    else:
        st.info("No leads yet. Add your first lead from the Leads page!")

    st.markdown('<div class="section-heading">📅 Upcoming Tasks Intelligence</div>', unsafe_allow_html=True)
    sorted_tasks = [t for t in tasks if not t.get("done") and t.get("due_date")]
    sorted_tasks = sorted(sorted_tasks, key=lambda x: x["due_date"])
    
    # Dashboard Pagination for Tasks
    DTR_PS = 5
    total_dtr = (len(sorted_tasks) // DTR_PS) + (1 if len(sorted_tasks) % DTR_PS > 0 else 0)
    if 'dash_tasks_page' not in st.session_state: st.session_state.dash_tasks_page = 1
    
    if total_dtr > 1:
        dtc1, dtc2, dtc3, dtc4 = st.columns([1,1,1,5])
        with dtc1:
            if st.button("⬅️", key="dtr_prev", disabled=st.session_state.dash_tasks_page <= 1):
                st.session_state.dash_tasks_page -= 1
                st.rerun()
        with dtc2:
            st.markdown(f"<div style='padding-top:10px; font-weight:700; color:#1b6656; font-size:0.8rem;'>{st.session_state.dash_tasks_page}/{total_dtr}</div>", unsafe_allow_html=True)
        with dtc3:
            if st.button("➡️", key="dtr_next", disabled=st.session_state.dash_tasks_page >= total_dtr):
                st.session_state.dash_tasks_page += 1
                st.rerun()
        
        s_idx = (st.session_state.dash_tasks_page - 1) * DTR_PS
        upcoming = sorted_tasks[s_idx:s_idx + DTR_PS]
    else:
        upcoming = sorted_tasks[:5]

    if upcoming:
        for t in upcoming:
            due = datetime.strptime(t["due_date"], "%Y-%m-%d").date()
            is_overdue = due < date.today()
            lname = get_lead_name(t.get("lead_id")) if t.get("lead_id") else "—"
            pcolor = {"High": "#ef4444", "Medium": "#fbbf24", "Low": "#10b981"}.get(t.get("priority","Medium"), "#ffffff")
            st.markdown(f"""
            <div class="glass-card" style="padding: 18px 24px; margin-bottom: 12px; border-left: 4px solid { '#ef4444' if is_overdue else 'var(--primary)' };">
                <div style="display:flex; justify-content:space-between; align-items:flex-start;">
                    <div>
                        <div style="font-family:'Outfit',sans-serif; font-size:1.05rem; font-weight:700; color:#1d4354; margin-bottom:4px;">{t['title']}</div>
                        <div style="color:#1b6656; font-size:0.78rem; font-weight:500;">
                            <span style="display:inline-flex; align-items:center; gap:5px;">👤 {lname}</span>
                        </div>
                    </div>
                    <div>
                        <span style="background: {pcolor}15; color: {pcolor}; padding: 4px 10px; border-radius: 6px; font-size: 0.65rem; font-weight: 800; border: 1px solid {pcolor}40; text-transform: uppercase;">
                            {t.get('priority','Medium')} Priority
                        </span>
                    </div>
                </div>
                <div style="margin-top:10px; display:flex; align-items:center; gap:15px; font-size:0.78rem;">
                    <span style="color:{ '#ef4444' if is_overdue else '#1b6656' }; font-weight:700;">
                        {'⚠️ OVERDUE' if is_overdue else '📅 DUE'} : {t['due_date']}
                    </span>
                    <span style="color:rgba(27, 102, 86, 0.4); margin-left:auto;">Action Required</span>
                </div>
            </div>
            """, unsafe_allow_html=True)
    else:
        st.info("No upcoming tasks!")


# ═══════════════════════════════════════════════════════════════════════════════
# SALES REPORT
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "💰 Sales Report":
    st.markdown("""
    <div class='page-header animate-in'>
        <div class='page-title'>Sales Intelligence</div>
        <div class='page-sub'>Excel-Style Ledger & Permanent Client Management</div>
    </div>
    """, unsafe_allow_html=True)

    # Defensive init
    if "recurring_clients" not in st.session_state:
        st.session_state.recurring_clients = db.get_recurring_clients()
    if "sales" not in st.session_state:
        st.session_state.sales = db.get_all_sales()

    # Initialize Recurring Clients Data if empty (Seed)
    if not st.session_state.recurring_clients:
        defaults = [
            ("DigiLlama - Monthly Bookkeeping", 45000, ""),
            ("H.S Power - Monthly Bookkeeping", 52000, ""),
            ("Saif Healthcare Limited - Tax Consultancy", 170000, ""),
            ("Lakki Broadcasting - Monthly Bookkeeping", 40000, ""),
            ("Inayat Medical Services - Monthly Bookkeeping", 30000, ""),
            ("The Kandle Co. - Monthly Bookkeeping", 40000, ""),
            ("Kulsum Medical City (KMC) - Monthly Bookkeeping", 63600, ""),
            ("Lone Star Restaurants", 132287, "31.5 hours x 15 = $472.5"),
            ("Olivier (Upwork)", 0, ""),
            ("HaVyn LifeStyle", 30000, "Referral from Bilal"),
            ("Zain Kazmi (Dasein)", 35000, "30k office + 5k rent agreement"),
            ("Hybrid Engineering SMC Pvt Limited", 15000, "Referral from Mustajab")
        ]
        for cl, am, nt in defaults:
            db.add_recurring_client({"client": cl, "default_amount": am, "default_notes": nt})
        st.session_state.recurring_clients = db.get_recurring_clients()

    tab_led, tab_bi, tab_perm = st.tabs(["📊 Monthly Ledger", "📈 Business Intelligence", "🏢 Permanent Clients"])

    with tab_perm:
        st.markdown('<div class="section-heading">Manage Permanent Clients</div>', unsafe_allow_html=True)
        st.markdown("Yeh clients har month ki default entry banna shuru ho jayenge jab aap naya month initialize karenge.")
        
        # Use simple form to add new permanent client
        with st.form("add_perm_client", clear_on_submit=True):
            pc1, pc2 = st.columns(2)
            new_pc_name = pc1.text_input("Client Name")
            new_pc_amount = pc2.number_input("Default Amount", min_value=0.0)
            new_pc_notes = st.text_input("Default Notes (e.g. Referral)")
            if st.form_submit_button("ADD PERMANENT CLIENT"):
                if new_pc_name:
                    db.add_recurring_client({"client": new_pc_name, "default_amount": new_pc_amount, "default_notes": new_pc_notes})
                    st.session_state.recurring_clients = db.get_recurring_clients()
                    st.success(f"Added {new_pc_name} to permanent list!")
                    st.rerun()

        # Display list with delete option
        for rc in st.session_state.recurring_clients:
            col1, col2, col3 = st.columns([3, 1, 0.5])
            col1.write(f"**{rc['client']}**")
            col2.write(f"PKR {rc['default_amount']:,.0f}")
            if col3.button("🗑️", key=f"del_rc_{rc['id']}"):
                db.delete_recurring_client(rc['id'])
                st.session_state.recurring_clients = db.get_recurring_clients()
                st.rerun()

    with tab_bi:
        st.markdown('<div class="section-heading">Revenue Intelligence & Trends</div>', unsafe_allow_html=True)
        
        all_sales_raw = db.get_all_sales()
        if all_sales_raw:
            all_years = [int(s["month_year"].split("-")[0]) for s in all_sales_raw]
            available_sales_years = sorted(list(set(all_years)), reverse=True)
            bi_year = st.selectbox("Intelligence Period (Year)", options=available_sales_years, index=0, key="bi_year_sel")
            
            df_all = pd.DataFrame(all_sales_raw)
            df_all['amount'] = pd.to_numeric(df_all['amount'], errors='coerce').fillna(0)
            df_all['dt'] = pd.to_datetime(df_all['month_year'], format='%Y-%m', errors='coerce')
            df_y = df_all[df_all['dt'].dt.year == bi_year]
            
            # 1. Monthly Revenue Trend
            monthly_trend = df_y[df_y['category'] == 'Confirmed'].groupby(df_y['dt'].dt.strftime('%m - %B'))['amount'].sum().reset_index()
            monthly_trend.columns = ['Month', 'Revenue']
            
            fig_trend = px.bar(monthly_trend, x='Month', y='Revenue', 
                              title=f"Monthly Confirmed Revenue - {bi_year}",
                              color='Revenue', color_continuous_scale='Mint',
                              text_auto='.2s')
            fig_trend.update_layout(template="plotly_white", font_family="Outfit", plot_bgcolor="rgba(0,0,0,0)")
            st.plotly_chart(fig_trend, use_container_width=True)
            
            c1, c2 = st.columns(2)
            with c1:
                # 2. Client Contribution (Top 10)
                client_cont = df_y[df_y['category'] == 'Confirmed'].groupby('client')['amount'].sum().sort_values(ascending=False).head(10).reset_index()
                
                # Truncate names for legend display to avoid layout cramping
                client_cont['display_name'] = client_cont['client'].apply(lambda x: x[:25] + '...' if len(x) > 28 else x)
                
                fig_client = px.pie(client_cont, values='amount', names='display_name', hole=0.5,
                                   title="Top 10 High-Value Clients",
                                   color_discrete_sequence=px.colors.qualitative.Prism)
                
                fig_client.update_traces(
                    textinfo='percent', 
                    textposition='outside',
                    hoverinfo='label+value+percent',
                    marker=dict(line=dict(color='#FFFFFF', width=2))
                )
                
                fig_client.update_layout(
                    template="plotly_white",
                    font_family="Outfit",
                    showlegend=True,
                    # Legend at bottom to give pie more width
                    legend=dict(
                        orientation="h",
                        yanchor="top",
                        y=-0.05,
                        xanchor="center",
                        x=0.5,
                        font=dict(size=10)
                    ),
                    margin=dict(t=60, l=60, r=60, b=100),
                    height=550,
                    title=dict(x=0.5, y=0.95, font=dict(size=18, weight=800, color="#1b6656"))
                )
                
                st.plotly_chart(fig_client, use_container_width=True, config={'displayModeBar': False})
                
            with c2:
                # 3. Revenue Breakdown - Stacked Bar (More readable than Sunburst)
                top_clients_all = df_y.groupby(['client', 'category'])['amount'].sum().reset_index()
                # Get the top 15 clients by total revenue to keep it clean
                top_15_names = df_y.groupby('client')['amount'].sum().sort_values(ascending=False).head(15).index
                df_top_15 = top_clients_all[top_clients_all['client'].isin(top_15_names)]
                
                fig_stat = px.bar(df_top_15, y='client', x='amount', color='category',
                                 title="Revenue Mix: Confirmed vs Potential (Top 15)",
                                 orientation='h',
                                 color_discrete_map={'Confirmed': '#1b6656', 'Potential': '#fbbf24'},
                                 category_orders={"client": top_15_names.tolist()}) # keep order
                
                fig_stat.update_layout(template="plotly_white", font_family="Outfit", 
                                      xaxis_title="Amount (PKR)", yaxis_title=None,
                                      legend_title=None, margin=dict(l=20, r=20, t=60, b=20))
                st.plotly_chart(fig_stat, use_container_width=True)
                
            # --- KPIs for the Year ---
            y_total = df_y[df_y['category'] == 'Confirmed']['amount'].sum()
            y_avg = df_y[df_y['category'] == 'Confirmed'].groupby(df_y['dt'].dt.month)['amount'].sum().mean() if not df_y.empty else 0
            y_pot = df_y[df_y['category'] == 'Potential']['amount'].sum()
            
            st.markdown(f"""
            <div class="kpi-container">
                <div class="kpi-card" style="--accent-gradient: linear-gradient(90deg, #1b6656, #7bb06b);">
                    <div class="kpi-icon">💰</div>
                    <div class="kpi-value" style="color:#1b6656">PKR {y_total:,.0f}</div>
                    <div class="kpi-label">{bi_year} Cumulative Revenue</div>
                </div>
                <div class="kpi-card" style="--accent-gradient: linear-gradient(90deg, #1d4354, #1b6656);">
                    <div class="kpi-icon">📈</div>
                    <div class="kpi-value" style="color:#1b6656">PKR {y_avg:,.0f}</div>
                    <div class="kpi-label">Monthly Average Yield</div>
                </div>
                <div class="kpi-card" style="--accent-gradient: linear-gradient(90deg, #fbbf24, #ef4444);">
                    <div class="kpi-icon">💎</div>
                    <div class="kpi-value" style="color:#ef4444">PKR {y_pot:,.0f}</div>
                    <div class="kpi-label">Untapped Potential</div>
                </div>
            </div>
            """, unsafe_allow_html=True)
        else:
            st.info("Insufficient sales data for Business Intelligence analysis.")

    with tab_led:
        # Selector for month and year
        st.markdown('<div class="luxury-container" style="padding:15px; margin-bottom:20px;">', unsafe_allow_html=True)
        lc1, lc2, lc3 = st.columns([1,1,1])
        
        # Determine available years/months from DB directly to ensure sync
        all_sales_raw = db.get_all_sales()
        if all_sales_raw:
            all_years = [int(s["month_year"].split("-")[0]) for s in all_sales_raw]
            distinct_years = sorted(list(set(all_years)), reverse=True)
        else:
            distinct_years = [datetime.now().year]
            
        sel_year = lc1.selectbox("Report Year", distinct_years, index=0)
        sel_month_name = lc2.selectbox("Report Month", 
                                     ["January", "February", "March", "April", "May", "June", 
                                      "July", "August", "September", "October", "November", "December"],
                                     index=datetime.now().month - 1)
        
        month_idx = ["January", "February", "March", "April", "May", "June", 
                     "July", "August", "September", "October", "November", "December"].index(sel_month_name) + 1
        c_month = f"{sel_year}-{month_idx:02d}"
        
        st.markdown('</div>', unsafe_allow_html=True)
        
        # Pull data for this specific month ONLY
        active_sales = [s for s in all_sales_raw if s["month_year"] == c_month]
        
        if not active_sales:
            st.warning(f"No ledger entries found for {sel_month_name} {sel_year}. Initialize now?")
            if st.button(f"🚀 INITIALIZE {sel_month_name} LEDGER WITH PERMANENT CLIENTS"):
                for rc in st.session_state.recurring_clients:
                    db.add_sale({
                        "month_year": c_month,
                        "category": "Confirmed",
                        "client": rc["client"],
                        "amount": rc["default_amount"],
                        "notes": rc["default_notes"],
                        "created_at": str(datetime.now())
                    })
                st.session_state.sales = db.get_all_sales()
                st.rerun()
        
        if active_sales:
            df_ledger = pd.DataFrame(active_sales)
            st.markdown(f"### <span style='color:#1b6656;'>{sel_month_name.upper()} {sel_year}</span> Executive Ledger", unsafe_allow_html=True)
            st.markdown("*Double-click values to edit. Ensure you click 'Save All Changes' to persistent storage.*")
            
            # Data Editor for Excel Feeling
            edited_ledger = st.data_editor(
                df_ledger[["client", "amount", "category", "notes", "id"]],
                use_container_width=True,
                num_rows="dynamic",
                column_config={
                    "category": st.column_config.SelectboxColumn("Status", options=["Confirmed", "Potential"], required=True),
                    "amount": st.column_config.NumberColumn("PKR Amount", format="PKR %d"),
                    "notes": st.column_config.TextColumn("Comments", width="large"),
                    "id": None # Hide ID
                },
                key=f"editor_{c_month}"
            )
            
            ec1, ec2 = st.columns([1,1])
            # Save logic
            if ec1.button("💾 SAVE ALL CHANGES", use_container_width=True):
                for s in active_sales:
                    db.delete_sale(s["id"])
                
                for _, row in edited_ledger.iterrows():
                    if row["client"]:
                        db.add_sale({
                            "month_year": c_month,
                            "category": row["category"] or "Confirmed",
                            "client": row["client"],
                            "amount": row["amount"] or 0,
                            "notes": row["notes"] or "",
                            "created_at": str(datetime.now())
                        })
                st.session_state.sales = db.get_all_sales()
                st.session_state.pending_note = {"msg": "Sales ledger successfully synchronized with ledger!", "title": "Sync Completed", "type": "success"}
                st.rerun()

            # Excel Export
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Prepare clean export data
                df_exp = edited_ledger[["client", "amount", "category", "notes"]].copy()
                df_exp.columns = ["🏆 CLIENT NAME", "💰 AMOUNT (PKR)", "🏷️ STATUS", "📝 ADDITIONAL NOTES"]
                
                # Start writing data from Row 3 to leave space for Title
                df_exp.to_excel(writer, index=False, sheet_name=f'Sales_{c_month}', startrow=2)
                
                # ─── PREMIUM STYLING ───
                workbook = writer.book
                worksheet = writer.sheets[f'Sales_{c_month}']
                
                # 1. ADD MAIN TITLE AT THE TOP
                report_title = f"EXECUTIVE SALES REPORT - {sel_month_name.upper()} {sel_year}"
                worksheet.merge_cells('A1:D1')
                title_cell = worksheet['A1']
                title_cell.value = report_title
                title_cell.font = Font(color='1B6656', bold=True, size=16)
                title_cell.alignment = Alignment(horizontal='center', vertical='center')
                title_cell.fill = PatternFill(start_color='E1F2EF', end_color='E1F2EF', fill_type='solid')
                worksheet.row_dimensions[1].height = 35 # Give it some height

                # Style Definitions
                header_bg = PatternFill(start_color='1B6656', end_color='1B6656', fill_type='solid') # Primary Green
                stripe_bg = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid') # Subtle Stripe
                
                header_font = Font(color='FFFFFF', bold=True, size=12)
                confirmed_font = Font(color='1B6656', bold=True)
                potential_font = Font(color='B45309', bold=True) # Amber/Gold
                
                alignment_center = Alignment(horizontal='center', vertical='center')
                alignment_left = Alignment(horizontal='left', vertical='center', indent=1)
                
                border_thin = Border(
                    left=Side(style='thin', color='CBD5E1'),
                    right=Side(style='thin', color='CBD5E1'),
                    top=Side(style='thin', color='CBD5E1'),
                    bottom=Side(style='thin', color='CBD5E1')
                )

                # Format Headers (Now on Row 3)
                for cell in worksheet[3]:
                    cell.fill = header_bg
                    cell.font = header_font
                    cell.alignment = alignment_center
                    cell.border = Border(bottom=Side(style='medium', color='7BB06B')) # Accent border

                # Apply Styles to Data Rows (Starting from Row 4)
                for row_idx, row in enumerate(worksheet.iter_rows(min_row=4, max_row=worksheet.max_row), start=4):
                    for cell in row:
                        cell.border = border_thin
                        cell.alignment = alignment_center
                        
                    # Column specific logic
                    worksheet.cell(row=row_idx, column=1).alignment = alignment_left
                    worksheet.cell(row=row_idx, column=2).number_format = '#,##0'
                    cat_cell = worksheet.cell(row=row_idx, column=3)
                    if cat_cell.value == 'Confirmed': cat_cell.font = confirmed_font
                    else: cat_cell.font = potential_font
                    worksheet.cell(row=row_idx, column=4).alignment = alignment_left

                    if row_idx % 2 == 0:
                        for cell in row: cell.fill = stripe_bg

                # Optimized Column Widths
                worksheet.column_dimensions['A'].width = 45
                worksheet.column_dimensions['B'].width = 20
                worksheet.column_dimensions['C'].width = 18
                worksheet.column_dimensions['D'].width = 50

                # ─── TOTAL SUMS AT THE BOTTOM ───
                last_row = worksheet.max_row + 2 
                
                c_sum = df_exp[df_exp["🏷️ STATUS"] == "Confirmed"]["💰 AMOUNT (PKR)"].sum()
                worksheet.cell(row=last_row, column=1, value="TOTAL CONFIRMED REVENUE").font = Font(bold=True, color='1B6656')
                worksheet.cell(row=last_row, column=2, value=c_sum).font = Font(bold=True, color='1B6656')
                worksheet.cell(row=last_row, column=2).number_format = '"PKR "#,##0'
                for col in range(1, 3):
                    worksheet.cell(row=last_row, column=col).fill = PatternFill(start_color='E1F2EF', end_color='E1F2EF', fill_type='solid')

                p_sum = df_exp[df_exp["🏷️ STATUS"] == "Potential"]["💰 AMOUNT (PKR)"].sum()
                worksheet.cell(row=last_row+1, column=1, value="TOTAL POTENTIAL PIPELINE").font = Font(bold=True, color='B45309')
                worksheet.cell(row=last_row+1, column=2, value=p_sum).font = Font(bold=True, color='B45309')
                worksheet.cell(row=last_row+1, column=2).number_format = '"PKR "#,##0'
                for col in range(1, 3):
                    worksheet.cell(row=last_row+1, column=col).fill = PatternFill(start_color='FFFBEB', end_color='FFFBEB', fill_type='solid')
                
            excel_data = output.getvalue()
            ec2.download_button(
                label="📊 DOWNLOAD EXCEL",
                data=excel_data,
                file_name=f"Sidekick_Sales_{c_month}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

            # Stats Bar
            c_rev = edited_ledger[edited_ledger["category"] == "Confirmed"]["amount"].sum()
            p_rev = edited_ledger[edited_ledger["category"] == "Potential"]["amount"].sum()
            
            st.markdown(f"""
            <div style="display:flex; justify-content:space-around; background:linear-gradient(90deg, #1b6656, #1d4354); color:white; padding:25px; border-radius:20px; margin-top:30px; border:1px solid rgba(255,255,255,0.1);">
                <div style="text-align:center;">
                    <div style="font-size:0.75rem; opacity:0.8; text-transform:uppercase; letter-spacing:0.1em;">Confirmed Revenue</div>
                    <div style="font-size:2rem; font-weight:800;">PKR {c_rev:,.0f}</div>
                </div>
                <div style="width:1px; background:rgba(255,255,255,0.2); margin:0 20px;"></div>
                <div style="text-align:center;">
                    <div style="font-size:0.75rem; opacity:0.8; text-transform:uppercase; letter-spacing:0.1em;">Potential Pipeline</div>
                    <div style="font-size:2rem; font-weight:800; color:#fbbf24;">PKR {p_rev:,.0f}</div>
                </div>
            </div>
            """, unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# LEADS
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "👥 Leads":
    st.markdown("""
    <div class='page-header animate-in'>
        <div class='page-title'>Leads Management</div>
        <div class='page-sub'>Prospect Portfolio & Pipeline Board</div>
    </div>
    """, unsafe_allow_html=True)

    tab1, tab2, tab3 = st.tabs(["🚀 Pipeline Board", "📋 List View", "➕ Add Lead"])
    
    with tab1:
        # Kanban style board
        k1, k2, k3 = st.columns(3)
        statuses = [
            ("New", "badge-new", k1, "🆕 New Opportunities"),
            ("In Progress", "badge-inprogress", k2, "⚡ Active Engagement"),
            ("Closed", "badge-closed", k3, "✅ Closed Deals")
        ]
        
        for status_name, badge_class, col, display_title in statuses:
            with col:
                st.markdown(f"""
                <div class="kanban-header">
                    <div class="kanban-title">{display_title}</div>
                    <div class="kanban-count">{sum(1 for l in leads if l.get("status") == status_name)}</div>
                </div>
                """, unsafe_allow_html=True)
                status_leads = [l for l in leads if l.get("status") == status_name]
                
                # Pagination
                KB_PS = 5
                total_kb = (len(status_leads) // KB_PS) + (1 if len(status_leads) % KB_PS > 0 else 0)
                kb_key = f"kb_l_{status_name.lower().replace(' ', '_')}"
                if kb_key not in st.session_state: st.session_state[kb_key] = 1
                
                if total_kb >= 1:
                    kc1, kc2, kc3 = st.columns([1,2,1])
                    with kc1:
                        if st.button("⬅️", key=f"p_{kb_key}", disabled=st.session_state[kb_key] <= 1, use_container_width=True):
                            st.session_state[kb_key] -= 1
                            st.rerun()
                    with kc2:
                        st.markdown(f"<p style='font-size:0.75rem; text-align:center; color:#1b6656; font-weight:800; padding-top:8px;'>{st.session_state[kb_key]} / {total_kb}</p>", unsafe_allow_html=True)
                    with kc3:
                        if st.button("➡️", key=f"n_{kb_key}", disabled=st.session_state[kb_key] >= total_kb, use_container_width=True):
                            st.session_state[kb_key] += 1
                            st.rerun()
                    
                    s_idx = (st.session_state[kb_key] - 1) * KB_PS
                    d_leads = status_leads[s_idx:s_idx + KB_PS]
                else:
                    d_leads = status_leads

                for l in d_leads:
                    sc = STATUS_COLORS.get(l.get("status", "New"), "badge-new")
                    tc = TEMP_COLORS.get(l.get("temperature", "Warm"), "badge-warm")
                    st.markdown(f"""
                    <div class="kanban-item animate-in">
                        <div style="font-weight:700; margin-bottom:6px; font-size:1.05rem;">{l['name']}</div>
                        <div style="font-size:0.75rem; color:var(--text-muted); margin-bottom:10px;">🏢 {l.get('company','N/A')}</div>
                        <div style="display:flex; justify-content:space-between; align-items:center;">
                            <span class="badge {tc}">{l.get('temperature','Warm')}</span>
                            <span style="font-size:0.65rem; color:rgba(255,255,255,0.3);">📅 {l.get('followup_date','-')}</span>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
                
                # Bottom Pagination for Kanban Column
                if total_kb >= 1:
                    b_kc1, b_kc2, b_kc3 = st.columns([1,2,1])
                    with b_kc1:
                        if st.button("⬅️", key=f"bp_{kb_key}", disabled=st.session_state[kb_key] <= 1, use_container_width=True):
                            st.session_state[kb_key] -= 1
                            st.rerun()
                    with b_kc2:
                        st.markdown(f"<p style='font-size:0.75rem; text-align:center; color:#1b6656; font-weight:800; padding-top:8px;'>{st.session_state[kb_key]} / {total_kb}</p>", unsafe_allow_html=True)
                    with b_kc3:
                        if st.button("➡️", key=f"bn_{kb_key}", disabled=st.session_state[kb_key] >= total_kb, use_container_width=True):
                            st.session_state[kb_key] += 1
                            st.rerun()

    with tab2:
        if not leads: st.info("Lead portfolio empty.")
        else:
            cf1, cf2, cf3 = st.columns([2, 1, 1])
            fsearch = cf1.text_input("🔍 Search Intelligence", placeholder="Name / Company / Email...", key="lead_search_main")
            fstatus = cf2.selectbox("Filter Status", ["All", "New", "In Progress", "Closed"], key="lead_filter_status")
            ftemp = cf3.selectbox("Filter Temperature", ["All", "Hot", "Warm", "Cold"], key="lead_filter_temp")
            filtered = leads
            if fsearch:
                f = fsearch.lower()
                filtered = [l for l in leads if f in l['name'].lower() or f in l.get('company','').lower() or f in l.get('email','').lower()]
            
            if fstatus != "All":
                filtered = [l for l in filtered if l.get("status") == fstatus]
            if ftemp != "All":
                filtered = [l for l in filtered if l.get("temperature") == ftemp]

            st.markdown(f"<div style='color:#1b6656; font-weight:800; font-size:0.9rem; margin-bottom:15px; text-transform:uppercase; letter-spacing:0.05em;'>📋 Pipeline Intelligence: {len(filtered)} Records Found</div>", unsafe_allow_html=True)

            # --- LEAD PAGINATION ENGINE ---
            LA_PS = 5
            total_p = (len(filtered) // LA_PS) + (1 if len(filtered) % LA_PS > 0 else 0)
            
            if 'lead_page' not in st.session_state: st.session_state.lead_page = 1
            if total_p > 0 and st.session_state.lead_page > total_p: st.session_state.lead_page = total_p
            
            if total_p >= 1:
                p_c1, p_c2, p_c3 = st.columns([1, 2, 1])
                with p_c1:
                    if st.button("⬅️", key="prev_lead", disabled=st.session_state.lead_page <= 1, use_container_width=True):
                        st.session_state.lead_page -= 1
                        st.rerun()
                with p_c2:
                    st.markdown(f"<p style='font-size:0.75rem; text-align:center; color:#1b6656; font-weight:800; padding-top:8px;'>{st.session_state.lead_page} / {total_p}</p>", unsafe_allow_html=True)
                with p_c3:
                    if st.button("➡️", key="next_lead", disabled=st.session_state.lead_page >= total_p, use_container_width=True):
                        st.session_state.lead_page += 1
                        st.rerun()
                
                s_i = (st.session_state.lead_page - 1) * LA_PS
                page_data = filtered[s_i:s_i + LA_PS]
            else:
                page_data = filtered
                st.session_state.lead_page = 1

            for l in page_data:
                sc = STATUS_COLORS.get(l.get("status", "New"), "badge-new")
                tc = TEMP_COLORS.get(l.get("temperature", "Warm"), "badge-warm")
                
                with st.expander(f"⚙️ MODIFY INTEL: {l['name']}", expanded=False):
                    with st.form(key=f"edit_lead_v2_{l['id']}"):
                        v1, v2 = st.columns(2)
                        u_name = v1.text_input("Lead Name", value=l['name'])
                        u_comp = v2.text_input("Company", value=l.get('company',''))
                        v3, v4 = st.columns(2)
                        u_email = v3.text_input("Email", value=l.get('email',''))
                        u_phone = v4.text_input("Phone", value=l.get('phone',''))
                        v5, v6, v7 = st.columns(3)
                        st_opts = ["New", "In Progress", "Closed"]
                        u_status = v5.selectbox("Status Update", st_opts, index=st_opts.index(l.get("status","New")) if l.get("status","New") in st_opts else 0)
                        tmp_opts = ["Hot", "Warm", "Cold"]
                        u_temp = v6.selectbox("Temp Shift", tmp_opts, index=tmp_opts.index(l.get("temperature","Warm")) if l.get("temperature","Warm") in tmp_opts else 0)
                        src_opts = ["Manual Entry", "Website", "Referral", "Ads", "Other"]
                        # Robust index lookup with guard
                        u_source = v7.selectbox("Channel", src_opts, index=src_opts.index(l.get("source","Manual Entry")) if l.get("source","Manual Entry") in src_opts else 0)
                        u_date = st.date_input("Next Milestone", value=datetime.strptime(l["followup_date"], "%Y-%m-%d").date() if l.get("followup_date") else date.today())
                        u_notes = st.text_area("Strategic Brief", value=l.get("notes",""))
                        u_commit, u_purge = st.columns(2)
                        if u_commit.form_submit_button("✅ SYNC CHANGES", use_container_width=True):
                            db.update_lead(l["id"], {"name": u_name, "company": u_comp, "email": u_email, "phone": u_phone, "status": u_status, "temperature": u_temp, "source": u_source, "followup_date": str(u_date), "notes": u_notes})
                            st.session_state.leads = db.get_all_leads()
                            st.session_state.pending_note = {"msg": "Lead intelligence updated successfully!", "title": "Update Success", "type": "success"}
                            st.rerun()
                        if u_purge.form_submit_button("🗑️ PURGE RECORD", use_container_width=True):
                            confirm_delete_lead(l['id'], l['name'])

                st.markdown(f"""
                <div class="glass-card" style="padding: 24px 30px; margin-bottom: 24px; border-left: 6px solid { '#1b6656' if l.get('status') == 'Closed' else '#7bb06b' }; box-shadow: 0 10px 30px -15px rgba(0,0,0,0.1);">
                    <div style="display:flex; justify-content:space-between; align-items:flex-start; margin-bottom:20px;">
                        <div>
                            <div style="font-family:'Outfit',sans-serif; font-size:1.3rem; font-weight:800; color:#1a2e26; letter-spacing:-0.01em;">{l['name']}</div>
                            <div style="color:#1b6656; font-size:0.9rem; font-weight:600; opacity:0.85; margin-top:2px;">🏢 {l.get('company','Private Entity')}</div>
                        </div>
                        <div style="display:flex; gap:12px;">
                            <span class="badge {sc}" style="padding: 6px 14px; font-size: 0.7rem;">{l.get('status','New')}</span>
                            <span class="badge {tc}" style="padding: 6px 14px; font-size: 0.7rem;">{l.get('temperature','Warm')}</span>
                        </div>
                    </div>
                    <div style="display:grid; grid-template-columns: 1.1fr 0.9fr; gap: 30px; font-size: 0.88rem; color:#475569;">
                        <div style="background:rgba(27,102,86,0.03); padding:18px; border-radius:14px; border:1px solid rgba(27,102,86,0.1);">
                            <div style="font-weight:800; color:#1b6656; text-transform:uppercase; font-size:0.7rem; margin-bottom:12px; letter-spacing:0.1em; opacity:0.7;">Communication Matrix</div>
                            <div style="margin-bottom:8px; display:flex; align-items:center; gap:8px;">📧 <a href="mailto:{l.get('email','')}" style="color:#1b6656; text-decoration:none; font-weight:500;">{l.get('email',l['id'])}</a></div>
                            <div style="display:flex; align-items:center; gap:8px;">📞 <span style="font-weight:500;">{l.get('phone','-')}</span></div>
                        </div>
                        <div style="background:rgba(29,67,84,0.03); padding:18px; border-radius:14px; border:1px solid rgba(29,67,84,0.1);">
                            <div style="font-weight:800; color:#1d4354; text-transform:uppercase; font-size:0.7rem; margin-bottom:12px; letter-spacing:0.1em; opacity:0.7;">Pipeline Context</div>
                            <div style="margin-bottom:8px; font-weight:700;">📅 Milestone: <span style="color:#1b6656">{l.get('followup_date','-')}</span></div>
                            <div style="opacity:0.6; font-size:0.75rem;">Source: {l.get('source','-')}</div>
                        </div>
                    </div>
                    {f'<div style="margin-top:20px; padding:18px; background:linear-gradient(135deg, #f8fafc, #f1f5f9); border-radius:14px; font-size:0.85rem; border-left:4px solid #1b6656; color:#334155; line-height:1.5;"><b>Intelligence Report:</b><br>{l["notes"]}</div>' if l.get("notes") else ''}
                </div>
                """, unsafe_allow_html=True)
            
            # --- BOTTOM PAGINATION FOR LEADS ---
            if total_p >= 1:
                st.markdown("<br>", unsafe_allow_html=True)
                blp_c1, blp_c2, blp_c3 = st.columns([1, 2, 1])
                with blp_c1:
                    if st.button("⬅️", key="prev_lead_bottom", disabled=st.session_state.lead_page <= 1, use_container_width=True):
                        st.session_state.lead_page -= 1
                        st.rerun()
                with blp_c2:
                    st.markdown(f"<p style='font-size:0.75rem; text-align:center; color:#1b6656; font-weight:800; padding-top:8px;'>{st.session_state.lead_page} / {total_p}</p>", unsafe_allow_html=True)
                with blp_c3:
                    if st.button("➡️", key="next_lead_bottom", disabled=st.session_state.lead_page >= total_p, use_container_width=True):
                        st.session_state.lead_page += 1
                        st.rerun()



            # Export
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("📥 Export to CSV"):
                df = pd.DataFrame(filtered)
                csv = df.to_csv(index=False)
                st.download_button("Download CSV", csv, "leads_export.csv", "text/csv")

    with tab3:
        st.markdown('<div class="section-heading">✦ Add New Lead</div>', unsafe_allow_html=True)
        with st.form("add_lead_form", clear_on_submit=True):
            st.markdown('<div class="luxury-container" style="padding:25px;">', unsafe_allow_html=True)
            c1, c2 = st.columns(2)
            name = c1.text_input("Lead Name *", placeholder="Individual or Company Name")
            company = c2.text_input("Company Name", placeholder="Organization details")
            
            c3, c4 = st.columns(2)
            email = c3.text_input("Email", placeholder="contact@example.com")
            phone = c4.text_input("Phone", placeholder="+1-555-0000")
            
            c5, c6, c7 = st.columns(3)
            status = c5.selectbox("Status", ["New", "In Progress", "Closed"])
            temp = c6.selectbox("Temperature", ["Hot", "Warm", "Cold"])
            source = c7.selectbox("Source", ["Manual Entry", "Website", "Referral", "Ads", "Other"])
            
            f_date = st.date_input("Follow-up Date", value=date.today() + timedelta(days=2))
            notes = st.text_area("Initial Summary", placeholder="Context or strategic notes...")
            st.markdown('</div>', unsafe_allow_html=True)
            submitted = st.form_submit_button("ADD LEAD TO PIPELINE", use_container_width=True)

            if submitted:
                if not name:
                    st.error("Lead Name is required!")
                else:
                    lead_data = {
                        "name": name, "company": company, "email": email, "phone": phone,
                        "status": status, "temperature": temp, "source": source, "notes": notes,
                        "followup_date": str(f_date), "created_at": str(datetime.now())
                    }
                    db.add_lead(lead_data)
                    st.session_state.leads = db.get_all_leads()
                    st.session_state.pending_note = {"msg": f"Lead {name} added to pipeline successfully!", "title": "Lead Added", "type": "success"}
                    st.rerun()


# ═══════════════════════════════════════════════════════════════════════════════
# TASKS
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "✅ Tasks":
    st.markdown("""
    <div class='page-header animate-in'>
        <div class='page-title'>Tasks Management</div>
        <div class='page-sub'>Execution & Milestone Tracking</div>
    </div>
    """, unsafe_allow_html=True)

    tab1, tab2, tab3 = st.tabs(["🚀 Pipeline Board", "📋 Task List", "➕ Add Task"])
    
    with tab1:
        k1, k2, k3 = st.columns(3)
        states = [
            ("Pending", "badge-new", k1, "🕒 To Do"),
            ("In Progress", "badge-inprogress", k2, "⚡ Doing"),
            ("Completed", "badge-closed", k3, "✅ Done")
        ]
        
        for state_name, badge_class, col, display_title in states:
            with col:
                st.markdown(f"""
                <div class="kanban-header">
                    <div class="kanban-title">{display_title}</div>
                    <div class="kanban-count">{sum(1 for t in tasks if t.get("status", "Pending") == state_name or (state_name=="Completed" and t.get("done")))}</div>
                </div>
                """, unsafe_allow_html=True)
                
                status_tasks = [t for t in tasks if t.get("status", "Pending") == state_name or (state_name=="Completed" and t.get("done"))]
                
                # Columnar Pagination for Task Kanban
                KT_PAGE_SIZE = 5
                total_kt_pages = (len(status_tasks) // KT_PAGE_SIZE) + (1 if len(status_tasks) % KT_PAGE_SIZE > 0 else 0)
                kt_state_key = f"kt_tasks_{state_name.lower().replace(' ', '_')}_page"
                if kt_state_key not in st.session_state: st.session_state[kt_state_key] = 1
                
                if total_kt_pages >= 1:
                    kt_c1, kt_c2, kt_c3 = st.columns([1,2,1])
                    with kt_c1:
                        if st.button("⬅️", key=f"btn_prev_{kt_state_key}", disabled=st.session_state[kt_state_key] <= 1, use_container_width=True):
                            st.session_state[kt_state_key] -= 1
                            st.rerun()
                    with kt_c2:
                        st.markdown(f"<p style='font-size:0.75rem; text-align:center; color:#1b6656; font-weight:800; padding-top:8px;'>{st.session_state[kt_state_key]} / {total_kt_pages}</p>", unsafe_allow_html=True)
                    with kt_c3:
                        if st.button("➡️", key=f"btn_next_{kt_state_key}", disabled=st.session_state[kt_state_key] >= total_kt_pages, use_container_width=True):
                            st.session_state[kt_state_key] += 1
                            st.rerun()
                    
                    start_kt = (st.session_state[kt_state_key] - 1) * KT_PAGE_SIZE
                    end_kt = start_kt + KT_PAGE_SIZE
                    display_tasks = status_tasks[start_kt:end_kt]
                else:
                    display_tasks = status_tasks

                for t in display_tasks:
                    sc = STATUS_COLORS.get(t.get("status", "Pending"), "badge-new")
                    pcolor = {"High": "#f43f5e", "Medium": "#fbbf24", "Low": "#10b981"}.get(t.get("priority","Medium"), "#ffffff")
                    st.markdown(f"""
                    <div class="kanban-item animate-in" style="border-left: 3px solid {pcolor};">
                        <div style="font-weight:700; margin-bottom:8px; font-size:1.05rem;">{t['title']}</div>
                        <div style="font-size:0.75rem; color:var(--text-muted); margin-bottom:12px;">👤 {get_lead_name(t.get('lead_id'))}</div>
                        <div style="display:flex; justify-content:space-between; align-items:center;">
                            <span class="badge {sc}">{t.get('status','Pending')}</span>
                            <span style="font-size:0.65rem; color:rgba(255,255,255,0.3);">📅 {t.get('due_date','-')}</span>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
                
                # Bottom Pagination for Task Kanban
                if total_kt_pages >= 1:
                    b_kt_c1, b_kt_c2, b_kt_c3 = st.columns([1,2,1])
                    with b_kt_c1:
                        if st.button("⬅️", key=f"btn_bprev_{kt_state_key}", disabled=st.session_state[kt_state_key] <= 1, use_container_width=True):
                            st.session_state[kt_state_key] -= 1
                            st.rerun()
                    with b_kt_c2:
                        st.markdown(f"<p style='font-size:0.75rem; text-align:center; color:#1b6656; font-weight:800; padding-top:8px;'>{st.session_state[kt_state_key]} / {total_kt_pages}</p>", unsafe_allow_html=True)
                    with b_kt_c3:
                        if st.button("➡️", key=f"btn_bnext_{kt_state_key}", disabled=st.session_state[kt_state_key] >= total_kt_pages, use_container_width=True):
                            st.session_state[kt_state_key] += 1
                            st.rerun()

    with tab3:
        st.markdown('<div class="section-heading">✦ Create New Task</div>', unsafe_allow_html=True)
        with st.form("add_task_form", clear_on_submit=True):
            st.markdown('<div class="luxury-container" style="padding:25px;">', unsafe_allow_html=True)
            title = st.text_input("Action Item Title *", placeholder="Critical follow-up regarding...")
            
            c1, c2 = st.columns(2)
            lead_options = ["— Independent Task —"] + [f"{l['name']} ({l.get('company','')})" for l in st.session_state.leads]
            lead_sel = c1.selectbox("Associate with Prospect", lead_options)
            priority = c2.selectbox("Operational Priority", ["High", "Medium", "Low"])
            
            c3, c4 = st.columns(2)
            due_date = c3.date_input("Execution Deadline", value=date.today() + timedelta(days=1))
            remind_email = c4.text_input("Reminder Email (Daily)", placeholder="email@example.com",
                                          value=st.session_state.settings.get("notify_email",""),
                                          help="Reminders will be sent daily until status is 'Completed'")
            
            description = st.text_area("Operational Details", placeholder="Comprehensive task breakdown...")
            t_status = st.selectbox("Current Lifecycle State", ["Pending", "In Progress", "Completed"])
            st.markdown('</div>', unsafe_allow_html=True)
            submitted = st.form_submit_button("DEPLOY TASK TO MATRIX", use_container_width=True)

            if submitted:
                if not title:
                    st.error("Title is required!")
                else:
                    lead_id = None
                    if lead_sel != "— Independent Task —":
                        idx = lead_options.index(lead_sel) - 1
                        lead_id = st.session_state.leads[idx]["id"]
                    task_data = {
                        "title": title, "lead_id": lead_id, "priority": priority,
                        "due_date": str(due_date), "description": description,
                        "remind_email": remind_email, "done": t_status == "Completed",
                        "created_at": str(datetime.now())
                    }
                    db.add_task(task_data)
                    st.session_state.tasks = db.get_all_tasks()
                    st.session_state.pending_note = {"msg": "New task deployed to matrix successfully!", "title": "Task Created", "type": "success"}
                    st.rerun()


    with tab2:
        # tasks handled globally
        if not tasks:
            st.info("No tasks yet. Create your first task!")
        else:
            cf1, cf2 = st.columns(2)
            fshow = cf1.selectbox("Show", ["All Tasks", "Open Only", "Done Only", "Overdue"])
            fprio = cf2.selectbox("Priority", ["All", "High", "Medium", "Low"])

            filtered = tasks[:]
            if fshow == "Open Only":
                filtered = [t for t in filtered if not t.get("done")]
            elif fshow == "Done Only":
                filtered = [t for t in filtered if t.get("done")]
            elif fshow == "Overdue":
                filtered = [t for t in filtered if not t.get("done") and t.get("due_date") 
                           and datetime.strptime(t["due_date"], "%Y-%m-%d").date() < date.today()]
            if fprio != "All":
                filtered = [t for t in filtered if t.get("priority") == fprio]

            # --- TASK PAGINATION ENGINE ---
            PAGE_SIZE_T = 5
            total_pages_t = (len(filtered) // PAGE_SIZE_T) + (1 if len(filtered) % PAGE_SIZE_T > 0 else 0)
            
            if 'task_page' not in st.session_state: st.session_state.task_page = 1
            if total_pages_t > 0 and st.session_state.task_page > total_pages_t: st.session_state.task_page = total_pages_t
            
            # Show pagination if there are any tasks (even 1 page) so user confirms it exists
            if total_pages_t >= 1:
                tp_c1, tp_c2, tp_c3 = st.columns([1, 2, 1])
                with tp_c1:
                    if st.button("⬅️", key="prev_task", disabled=st.session_state.task_page <= 1, use_container_width=True):
                        st.session_state.task_page -= 1
                        st.rerun()
                with tp_c2:
                    st.markdown(f"<p style='font-size:0.75rem; text-align:center; color:#1b6656; font-weight:800; padding-top:8px;'>{st.session_state.task_page} / {total_pages_t}</p>", unsafe_allow_html=True)
                with tp_c3:
                    if st.button("➡️", key="next_task", disabled=st.session_state.task_page >= total_pages_t, use_container_width=True):
                        st.session_state.task_page += 1
                        st.rerun()
                
                start_idx_t = (st.session_state.task_page - 1) * PAGE_SIZE_T
                end_idx_t = start_idx_t + PAGE_SIZE_T
                page_data_t = filtered[start_idx_t:end_idx_t]
            else:
                page_data_t = filtered
                st.session_state.task_page = 1

            for t in page_data_t:
                done = t.get("done", False)
                due = datetime.strptime(t["due_date"], "%Y-%m-%d").date() if t.get("due_date") else None
                is_overdue = due and due < date.today() and not done
                css = "done" if done else ("overdue" if is_overdue else "")
                pcolor = {"High": "#ff6584", "Medium": "#ffb347", "Low": "#43e97b"}.get(t.get("priority","Medium"), "#ffffff")
                lname = get_lead_name(t.get("lead_id")) if t.get("lead_id") else "—"

                with st.expander(f"**{t['title']}**", expanded=False):
                    with st.form(key=f"edit_task_{t['id']}"):
                        e_title = st.text_input("Task Title", value=t['title'])
                        ec1, ec2 = st.columns(2)
                        
                        # Lead mapping for selection
                        l_opts = ["— Independent Task —"] + [f"{l['name']} ({l.get('company','')})" for l in st.session_state.leads]
                        current_l_idx = 0
                        if t.get("lead_id"):
                            for idx, lead in enumerate(st.session_state.leads):
                                if lead["id"] == t["lead_id"]:
                                    current_l_idx = idx + 1
                                    break
                        
                        e_lead = ec1.selectbox("Link to Lead", l_opts, index=current_l_idx)
                        prio_opts = ["High", "Medium", "Low"]
                        e_prio = ec2.selectbox("Priority", prio_opts, index=prio_opts.index(t.get("priority","Medium")) if t.get("priority","Medium") in prio_opts else 1)
                        
                        ec3, ec4, ec5 = st.columns(3)
                        e_due = ec3.date_input("Due Date", value=datetime.strptime(t["due_date"], "%Y-%m-%d").date() if t.get("due_date") else date.today())
                        e_remind = ec4.text_input("Reminder Email (Daily)", value=t.get("remind_email",""))
                        stat_opts = ["Pending", "In Progress", "Completed"]
                        cur_stat = t.get("status", "Completed" if t.get("done") else "Pending")
                        e_status = ec5.selectbox("Status", stat_opts, index=stat_opts.index(cur_stat) if cur_stat in stat_opts else 0)
                        e_desc = st.text_area("Description", value=t.get("description",""))
                        
                        if st.form_submit_button("💾 Save Changes", use_container_width=True):
                            lead_id = None
                            if e_lead != "— No lead —":
                                l_idx = l_opts.index(e_lead) - 1
                                lead_id = st.session_state.leads[l_idx]["id"]
                            
                            upd_task = {
                                "title": e_title,
                                "lead_id": lead_id,
                                "priority": e_prio,
                                "due_date": str(e_due),
                                "remind_email": e_remind,
                                "done": e_status == "Completed",
                                "description": e_desc
                            }
                            db.update_task(t["id"], upd_task)
                            st.session_state.tasks = db.get_all_tasks()
                            st.session_state.pending_note = {"msg": "Task protocol successfully updated!", "title": "Update: Success", "type": "success"}
                            st.rerun()

                col1, col2, col3 = st.columns([8, 1, 1])
                with col1:
                    st.markdown(f"""
                    <div class="glass-card {'animate-in' if not done else ''}" style="padding: 15px 20px; border-left: 4px solid { '#10b981' if done else ('#ef4444' if is_overdue else pcolor) }; opacity: {0.6 if done else 1};">
                        <div style="display:flex; align-items:center; justify-content:space-between;">
                            <div style="display:flex; align-items:center; gap:12px;">
                                <div style="width:10px; height:10px; border-radius:50%; background:{pcolor};"></div>
                                <b style="font-family:'Outfit',sans-serif; font-size:1rem; {'text-decoration:line-through; color:rgba(27,102,86,0.3)' if done else 'color:#1a2e26'}">{t['title']}</b>
                            </div>
                            <span class="badge { 'badge-inprogress' if t.get('status') == 'In Progress' else ('badge-new' if t.get('status') == 'Pending' else 'badge-closed') }">
                                {t.get('status', 'Completed' if done else 'Pending')}
                            </span>
                        </div>
                        <div style="margin-top:10px; font-size:0.78rem; color:#1b6656; display:flex; gap:15px;">
                            <span>👤 {get_lead_name(t.get("lead_id"), l_map)}</span>
                            <span style="color:{ '#ef4444' if is_overdue else '#1b6656' }; font-weight:600;">{'⚠️' if is_overdue else '📅'} {t.get('due_date','-')}</span>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
                with col2:
                    if st.button("🔄" if done else "✓", key=f"toggle_{t['id']}", use_container_width=True):
                        new_done = not t.get('done', False)
                        new_status = "Completed" if new_done else "Pending"
                        db.update_task(t['id'], {"done": new_done, "status": new_status})
                        st.session_state.tasks = db.get_all_tasks()
                        st.rerun()
                with col3:
                    if st.button("✕", key=f"deltask_{t['id']}", use_container_width=True):
                        db.delete_task(t['id'])
                        st.session_state.tasks = db.get_all_tasks()
                        st.session_state.pending_note = {"msg": "Task purged from matrix.", "title": "Protocol: Deleted", "type": "warning"}
                        st.rerun()

            # --- BOTTOM PAGINATION ---
            if total_pages_t >= 1:
                st.markdown("<br>", unsafe_allow_html=True)
                btp_c1, btp_c2, btp_c3 = st.columns([1, 2, 1])
                with btp_c1:
                    if st.button("⬅️", key="prev_task_bottom", disabled=st.session_state.task_page <= 1, use_container_width=True):
                        st.session_state.task_page -= 1
                        st.rerun()
                with btp_c2:
                    st.markdown(f"<p style='font-size:0.75rem; text-align:center; color:#1b6656; font-weight:800; padding-top:8px;'>{st.session_state.task_page} / {total_pages_t}</p>", unsafe_allow_html=True)
                with btp_c3:
                    if st.button("➡️", key="next_task_bottom", disabled=st.session_state.task_page >= total_pages_t, use_container_width=True):
                        st.session_state.task_page += 1
                        st.rerun()


# ═══════════════════════════════════════════════════════════════════════════════
# REMINDERS
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "📧 Reminders":
    st.markdown("""
    <div class='page-header animate-in'>
        <div class='page-title'>Communications Relay</div>
        <div class='page-sub'>Automated Stakeholder Engagement Control</div>
    </div>
    """, unsafe_allow_html=True)

    # leads and tasks are already globally filtered by the time engine
    today = date.today()

    # Dynamic period header
    st.markdown(f"""
        <div style="margin-bottom:20px; font-size:0.85rem; color:#1b6656; font-weight:700; opacity:0.8; text-align:right;">
            🕒 Monitoring Period: <span style="color:#1d4354">{selected_year}</span> {f' - <span style="color:#1d4354">{selected_month_name}</span>' if selected_month > 0 else ''}
        </div>
    """, unsafe_allow_html=True)

    # Task Filtering
    overdue_t = [t for t in tasks if not t.get("done") and t.get("due_date") 
               and datetime.strptime(t["due_date"], "%Y-%m-%d").date() < today]
    due_today_t = [t for t in tasks if not t.get("done") and t.get("due_date") == str(today)]
    due_soon_t = [t for t in tasks if not t.get("done") and t.get("due_date") 
                and today < datetime.strptime(t["due_date"], "%Y-%m-%d").date() <= today + timedelta(days=3)]

    # Lead Filtering
    overdue_l = [l for l in leads if l.get("status") != "Closed" and l.get("followup_date") 
               and datetime.strptime(l["followup_date"], "%Y-%m-%d").date() < today]
    due_today_l = [l for l in leads if l.get("status") != "Closed" and l.get("followup_date") == str(today)]
    due_soon_l = [l for l in leads if l.get("status") != "Closed" and l.get("followup_date") 
                and today < datetime.strptime(l["followup_date"], "%Y-%m-%d").date() <= today + timedelta(days=3)]

    st.markdown(f"""
    <div style="display:grid; grid-template-columns: 1fr 1fr 1fr; gap: 20px; margin-bottom: 30px;">
        <div class="luxury-container" style="text-align:center; padding: 20px; background: rgba(239, 68, 68, 0.05); border-color: rgba(239, 68, 68, 0.2);">
            <div style="font-size: 0.7rem; color: #ef4444; font-weight:800; text-transform: uppercase; letter-spacing:0.1em;">Critical Overdue</div>
            <div style="font-size: 1.8rem; font-weight: 800; color:#1d4354;">{len(overdue_t) + len(overdue_l)}</div>
        </div>
        <div class="luxury-container" style="text-align:center; padding: 20px; background: rgba(27, 102, 86, 0.05); border-color: rgba(27, 102, 86, 0.2);">
            <div style="font-size: 0.7rem; color: #1b6656; font-weight:800; text-transform: uppercase; letter-spacing:0.1em;">Immediate Focus</div>
            <div style="font-size: 1.8rem; font-weight: 800; color:#1d4354;">{len(due_today_t) + len(due_today_l)}</div>
        </div>
        <div class="luxury-container" style="text-align:center; padding: 20px; background: rgba(123, 176, 107, 0.05); border-color: rgba(123, 176, 107, 0.2);">
            <div style="font-size: 0.7rem; color: #7bb06b; font-weight:800; text-transform: uppercase; letter-spacing:0.1em;">72h Forecast</div>
            <div style="font-size: 1.8rem; font-weight: 800; color:#1d4354;">{len(due_soon_t) + len(due_soon_l)}</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<div class="section-heading">📤 Global Communication Relay</div>', unsafe_allow_html=True)
    with st.form("reminder_form"):
        st.markdown('<div class="luxury-container" style="padding:25px;">', unsafe_allow_html=True)
        to_email = st.text_input("Recipient Dispatch Email *", value=st.session_state.settings.get("notify_email",""))
        subject = st.text_input("Operational Brief Subject", value=f"Enterprise Status Report — {today}")
        
        st.markdown('<div style="color:var(--text-muted); font-size:0.75rem; margin-top:15px; margin-bottom:10px; text-transform:uppercase; font-weight:700; letter-spacing:0.05em;">Data Selection Protocols</div>', unsafe_allow_html=True)
        c_t1, c_t2, c_t3 = st.columns(3)
        include_overdue_t = c_t1.checkbox("Overdue Matrix", value=True, help="Wo Tasks jin ki date guzar chuki hai (Pending Tasks)")
        include_today_t = c_t2.checkbox("Daily Active Tasks", value=True, help="Wo Tasks jo aaj complete karne hain")
        include_soon_t = c_t3.checkbox("72h Task Forecast", value=True, help="Agle 3 dino mein due hone wale Tasks")

        c_l1, c_l2, c_l3 = st.columns(3)
        include_overdue_l = c_l1.checkbox("Stalled Portfolio", value=True, help="Wo Leads jin ka follow-up late ho chuka hai")
        include_today_l = c_l2.checkbox("Daily Active Leads", value=True, help="Wo Leads jin se aaj rabta karna hai")
        include_soon_l = c_l3.checkbox("72h Portfolio Forecast", value=True, help="Agle 3 dino mein rabta talab Leads")

        custom_msg = st.text_area("Executive Summary Supplement", placeholder="Add high-level context or personal directives...")
        st.markdown('</div>', unsafe_allow_html=True)

        if st.form_submit_button("INITIATE GLOBAL DISPATCH", use_container_width=True):
            if not to_email:
                st.error("Please enter an email address!")
            else:
                sections = ""
                
                def task_html(t_list, title, color):
                    if not t_list:
                        return ""
                    rows = "".join([f"<tr><td style='padding:6px 10px'>{t['title']}</td><td style='padding:6px 10px'>{get_lead_name(t.get('lead_id'))}</td><td style='padding:6px 10px'>{t.get('due_date','-')}</td><td style='padding:6px 10px'>{t.get('priority','Medium')}</td></tr>" for t in t_list])
                    return f"""
                    <h3 style='color:{color};font-family:sans-serif'>{title}</h3>
                    <table style='border-collapse:collapse;width:100%;font-family:sans-serif;font-size:14px;border:1px solid #ddd'>
                    <tr style='background:#f9f9f9;text-align:left'><th style='padding:8px 10px'>Task</th><th style='padding:8px 10px'>Lead</th><th style='padding:8px 10px'>Due</th><th style='padding:8px 10px'>Priority</th></tr>
                    {rows}
                    </table><br>
                    """

                def lead_html(l_list, title, color):
                    if not l_list:
                        return ""
                    rows = "".join([f"<tr><td style='padding:6px 10px'>{l['name']}</td><td style='padding:6px 10px'>{l.get('company','-')}</td><td style='padding:6px 10px'>{l.get('followup_date','-')}</td><td style='padding:6px 10px'>{l.get('temperature','-')}</td></tr>" for l in l_list])
                    return f"""
                    <h3 style='color:{color};font-family:sans-serif'>{title}</h3>
                    <table style='border-collapse:collapse;width:100%;font-family:sans-serif;font-size:14px;border:1px solid #ddd'>
                    <tr style='background:#f9f9f9;text-align:left'><th style='padding:8px 10px'>Lead Name</th><th style='padding:8px 10px'>Company</th><th style='padding:8px 10px'>Follow-up</th><th style='padding:8px 10px'>Temp</th></tr>
                    {rows}
                    </table><br>
                    """

                if include_overdue_t: sections += task_html(overdue_t, "⚠️ Overdue Tasks", "#e53935")
                if include_today_t: sections += task_html(due_today_t, "📅 Tasks Due Today", "#f57c00")
                if include_soon_t: sections += task_html(due_soon_t, "🔔 Tasks Due in 3 Days", "#1976d2")
                
                if include_overdue_l: sections += lead_html(overdue_l, "⚠️ Overdue Lead Follow-ups", "#e53935")
                if include_today_l: sections += lead_html(due_today_l, "📅 Lead Follow-ups Today", "#f57c00")
                if include_soon_l: sections += lead_html(due_soon_l, "🔔 Lead Follow-ups in 3 Days", "#1976d2")

                body = f"""
                <div style='font-family:sans-serif;max-width:600px;margin:0 auto'>
                    <h1 style='color:#1b6656'>🎯 Sidekick Reminders</h1>
                    <p style='color:#666'>{today.strftime('%A, %B %d, %Y')}</p>
                    {f'<div style="background:#f0f8f6;padding:15px;border-radius:10px;margin-bottom:20px">{custom_msg}</div>' if custom_msg else ''}
                    <hr style='border:none;border-top:1px solid #eee;margin:20px 0'>
                    {sections or '<p>No items to report!</p>'}
                    <hr style='border:none;border-top:1px solid #eee;margin:20px 0'>
                    <p style='color:#999;font-size:12px'>Sent from Sidekick Tasks</p>
                </div>
                """
                ok, msg = send_email(subject, body, to_email)
                if ok:
                    st.success(f"✅ Reminder email sent to {to_email}!")
                else:
                    st.error(f"Failed to send: {msg}. Check Settings → SMTP config.")

    # Show lists with Pagination
    st.markdown("<br>", unsafe_allow_html=True)
    
    col_v1, col_v2 = st.columns(2)
    with col_v1:
        st.markdown('<h4 style="font-size:0.9rem; color:#ef4444; margin-bottom:15px; text-transform:uppercase; letter-spacing:0.1em; font-weight:800;">⚠️ Critical Task Overdue</h4>', unsafe_allow_html=True)
        if overdue_t:
            PS_OT = 5
            total_pot = (len(overdue_t) // PS_OT) + (1 if len(overdue_t) % PS_OT > 0 else 0)
            if 'ot_page' not in st.session_state: st.session_state.ot_page = 1
            if st.session_state.ot_page > total_pot: st.session_state.ot_page = total_pot

            if total_pot > 1:
                ot_c1, ot_c2, ot_c3 = st.columns([1,1,1])
                with ot_c1: 
                    if st.button("⬅️", key="ot_prev", disabled=st.session_state.ot_page <= 1):
                        st.session_state.ot_page -= 1
                        st.rerun()
                with ot_c2: st.markdown(f"<p style='text-align:center;font-size:0.7rem;margin-top:10px;'>{st.session_state.ot_page}/{total_pot}</p>", unsafe_allow_html=True)
                with ot_c3:
                    if st.button("➡️", key="ot_next", disabled=st.session_state.ot_page >= total_pot):
                        st.session_state.ot_page += 1
                        st.rerun()
                
                os_idx = (st.session_state.ot_page - 1) * PS_OT
                disp_ot = overdue_t[os_idx:os_idx+PS_OT]
            else:
                disp_ot = overdue_t

            for t in disp_ot:
                st.markdown(f"""
                <div class="glass-card animate-in" style="padding:15px; border-left:3px solid #ef4444; margin-bottom:10px;">
                    <div style="font-weight:700; color:#1d4354; font-family:'Outfit',sans-serif;">{t['title']}</div>
                    <div style="font-size:0.75rem; color:#475569; margin-top:5px; display:flex; justify-content:space-between;">
                        <span>Expired: {t['due_date']}</span>
                        <span>Prospect: {get_lead_name(t.get('lead_id'))}</span>
                    </div>
                </div>
                """, unsafe_allow_html=True)
        else:
            st.info("No tasks overdue in this period.")
                
    with col_v2:
        st.markdown('<h4 style="font-size:0.9rem; color:#f59e0b; margin-bottom:15px; text-transform:uppercase; letter-spacing:0.1em; font-weight:800;">👤 Stalled Lead Portfolio</h4>', unsafe_allow_html=True)
        if overdue_l:
            PS_OL = 5
            total_pol = (len(overdue_l) // PS_OL) + (1 if len(overdue_l) % PS_OL > 0 else 0)
            if 'ol_page' not in st.session_state: st.session_state.ol_page = 1
            if st.session_state.ol_page > total_pol: st.session_state.ol_page = total_pol

            if total_pol > 1:
                ol_c1, ol_c2, ol_c3 = st.columns([1,1,1])
                with ol_c1: 
                    if st.button("⬅️", key="ol_prev", disabled=st.session_state.ol_page <= 1):
                        st.session_state.ol_page -= 1
                        st.rerun()
                with ol_c2: st.markdown(f"<p style='text-align:center;font-size:0.7rem;margin-top:10px;'>{st.session_state.ol_page}/{total_pol}</p>", unsafe_allow_html=True)
                with ol_c3:
                    if st.button("➡️", key="ol_next", disabled=st.session_state.ol_page >= total_pol):
                        st.session_state.ol_page += 1
                        st.rerun()
                
                ols_idx = (st.session_state.ol_page - 1) * PS_OL
                disp_ol = overdue_l[ols_idx:ols_idx+PS_OL]
            else:
                disp_ol = overdue_l

            for l in disp_ol:
                st.markdown(f"""
                <div class="glass-card animate-in" style="padding:15px; border-left:3px solid #f59e0b; margin-bottom:10px;">
                    <div style="font-weight:700; color:#1d4354; font-family:'Outfit',sans-serif;">{l['name']}</div>
                    <div style="font-size:0.75rem; color:#475569; margin-top:5px; display:flex; justify-content:space-between;">
                        <span>Last Threshold: {l['followup_date']}</span>
                        <span>Entity: {l.get('company','-')}</span>
                    </div>
                </div>
                """, unsafe_allow_html=True)
        else:
            st.info("No lead follow-ups overdue in this period.")

    # Automated Logs
    st.markdown('<div class="section-heading">🤖 Automated Runs (History)</div>', unsafe_allow_html=True)
    logs = db.get_logs(limit=5)
    if logs:
        for log in logs:
            st.markdown(f"""
            <div style="background:rgba(255,255,255,0.03); border:1px solid rgba(255,255,255,0.1); border-radius:8px; padding:8px 12px; margin-bottom:6px; font-size:0.8rem;">
                <b>📅 Run:</b> {log['timestamp'][:16]} | 
                <b>📧 Sent:</b> {log['emails_sent']} emails | 
                <b>✅ Tasks Chk:</b> {log.get('tasks_checked', 0)} | 
                <b>👥 Leads Chk:</b> {log.get('leads_checked', 0)}
            </div>
            """, unsafe_allow_html=True)
    else:
        st.info("No automated runs logged yet.")


# SETTINGS
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "⚙️ Settings":
    st.markdown("""
    <div class='page-header animate-in'>
        <div class='page-title'>System Configuration</div>
        <div class='page-sub'>Enterprise Infrastructure & Security Protocols</div>
    </div>
    """, unsafe_allow_html=True)

    s = st.session_state.settings
    with st.form("settings_form"):
        st.markdown('<div class="luxury-container" style="padding:25px;">', unsafe_allow_html=True)
        st.markdown('<div style="color:var(--primary); font-weight:700; margin-bottom:15px; text-transform:uppercase; font-size:0.75rem; letter-spacing:0.1em;">📧 SMTP Communication Infrastructure</div>', unsafe_allow_html=True)
        st.info("System Protocol: Use smtp.gmail.com (Port 587) with App Passwords for secure relay.")
        
        c1, c2 = st.columns(2)
        smtp_host = c1.text_input("Gateway Host", value=s.get("smtp_host","smtp.gmail.com"))
        smtp_port = c2.text_input("Access Port", value=str(s.get("smtp_port",587)))
        smtp_user = st.text_input("Authenticated Identity (Sender)", value=s.get("smtp_user",""), placeholder="enterprise@domain.com")
        smtp_pass = st.text_input("Encryption Access Token (App Pass)", value=s.get("smtp_pass",""), type="password")
        
        st.markdown('<hr style="border:none; border-top:1px solid rgba(255,255,255,0.05); margin:20px 0;">', unsafe_allow_html=True)
        st.markdown('<div style="color:var(--secondary); font-weight:700; margin-bottom:15px; text-transform:uppercase; font-size:0.75rem; letter-spacing:0.1em;">🤖 Artificial Intelligence & Automation</div>', unsafe_allow_html=True)
        
        notify_email = st.text_input("Operational Alert Relay", value=s.get("notify_email",""), placeholder="security@domain.com")
        gemini_api_key = st.text_input("Neural Core API Key (Gemini)", value=s.get("gemini_api_key",""), type="password")
        
        auto_reminders = st.toggle("Activate Autonomous Daily Briefing", value=s.get("auto_reminders", True))
        st.markdown('</div>', unsafe_allow_html=True)
        
        if st.form_submit_button("SYNCHRONIZE SYSTEM ARCHITECTURE", use_container_width=True):
            st.session_state.settings = {
                "smtp_host": smtp_host,
                "smtp_port": int(smtp_port) if smtp_port.isdigit() else 587,
                "smtp_user": smtp_user,
                "smtp_pass": smtp_pass,
                "notify_email": notify_email,
                "gemini_api_key": gemini_api_key,
                "auto_reminders": auto_reminders,
                "last_auto_run": s.get("last_auto_run", "")
            }
            db.save_settings(st.session_state.settings)
            st.success("Core Configuration Synchronized!")

    st.markdown('<div class="section-heading">🧪 Protocol Verification</div>', unsafe_allow_html=True)
    with st.form("test_email"):
        st.markdown('<div class="luxury-container" style="padding:20px;">', unsafe_allow_html=True)
        test_to = st.text_input("Target Relay for Verification", value=s.get("notify_email",""))
        st.markdown('</div>', unsafe_allow_html=True)
        if st.form_submit_button("EXECUTE RELAY TEST"):
            ok, msg = send_email(
                "Sidekick Protocol Verification ⚡",
                "<h2>Communication Established</h2><p>Enterprise relay configuration verified successfully.</p>",
                test_to
            )
            if ok:
                st.success("Relay test completed successfully!")
            else:
                st.error(f"Relay Verification Failed: {msg}")

    # Logout & Data
    st.markdown('<div class="section-heading">🛠️ Infrastructure Maintenance</div>', unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1:
        if st.button("TERMINATE SESSION", key="logout_settings", use_container_width=True):
            st.session_state.authenticated = False
            st.rerun()
    with c2:
        if st.button("PURGE TEMPORARY CACHE", use_container_width=True):
            st.warning("Session memory cleared. Physical sectors remain intact.")
            st.rerun()

    st.markdown('<div class="section-heading">🗄️ Database Governance</div>', unsafe_allow_html=True)
    col1, col2 = st.columns(2)
    with col1:
        if st.button("SECURE ERASE: LEAD PORTFOLIO", use_container_width=True):
            # This should actually call db.delete all or similar
            # For safety, let's not implement a mass delete here yet, but fix the call
            st.warning("Feature deactivated for safety. Use individual purge.")
    with col2:
        if st.button("SECURE ERASE: OPERATIONS MATRIX", use_container_width=True):
            st.warning("Feature deactivated for safety. Use individual purge.")

# EMAIL MARKETING HUB
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "📢 Email Marketing":
    st.markdown("""
    <div class='page-header animate-in'>
        <div class='page-title'>Email Marketing Command</div>
        <div class='page-sub'>Multi-Channel Campaign Architecture & Outreach</div>
    </div>
    """, unsafe_allow_html=True)

    tab1, tab2, tab3 = st.tabs(["📥 Bulk Dispatch", "📄 Templates", "📊 Campaign History"])

    with tab2:
        st.markdown('<div class="section-heading">✦ Template Management</div>', unsafe_allow_html=True)
        
        # Initialize state for confirm/edit
        if "delete_confirm_id" not in st.session_state:
            st.session_state.delete_confirm_id = None
        if "edit_template_id" not in st.session_state:
            st.session_state.edit_template_id = None

        templates = db.get_all_templates()
        edit_id = st.session_state.edit_template_id
        edit_data = next((t for t in templates if t['id'] == edit_id), None) if edit_id else None

        initial_body = edit_data['body'] if edit_data else ""

        # Aggressively unwrap HTML to show ONLY plain text for editing
        if initial_body and (initial_body.strip().startswith("<") or "<html>" in initial_body.lower()):
            # 1. Try to extract from our premium content div first
            content_match = re.search(r'<div class="content">(.*?)</div>', initial_body, re.DOTALL)
            if content_match:
                body_content = content_match.group(1)
            else:
                body_content = initial_body
            
            # 2. Clean up: convert <br> to \n and strip ALL other html tags
            clean_text = body_content.replace("<br>", "\n").replace("<BR>", "\n")
            clean_text = re.sub(r'<[^>]*>', '', clean_text) # Strip all remaining tags
            initial_body = clean_text.strip()

        # --- Template Creation/Edit Form ---
        with st.form("template_form"):
            st.markdown('<div class="luxury-container" style="padding:25px;">', unsafe_allow_html=True)
            if edit_id:
                st.markdown(f"#### 📝 Editing: <span style='color:var(--primary);'>{edit_data['name'] if edit_data else 'Template'}</span>", unsafe_allow_html=True)
            else:
                st.markdown("#### 📄 Create New Template", unsafe_allow_html=True)
            
            t_name = st.text_input("Template Display Name", value=edit_data['name'] if edit_data else "", placeholder="e.g., Welcome Series")
            t_subject = st.text_input("Email Subject Line", value=edit_data['subject'] if edit_data else "", placeholder="Professional Subject Line")
            
            # Gemini AI Integration for Writing
            st.markdown('<div style="margin-top:15px; margin-bottom:5px; font-size:0.75rem; color:var(--primary); font-weight:800; text-transform:uppercase;">Brainstorm with Gemini AI</div>', unsafe_allow_html=True)
            ai_prompt = st.text_input("Briefly tell AI what to write...", placeholder="Write a warm greeting to a new lead who just signed up", label_visibility="collapsed")
            if st.form_submit_button("✨ GENERATE AI CONTENT", use_container_width=True):
                if not st.session_state.settings.get("gemini_api_key"):
                    st.error("Please configure Gemini API Key in Settings.")
                elif not ai_prompt:
                    st.warning("Please tell AI what to write first!")
                else:
                    try:
                        import google.generativeai as genai
                        genai.configure(api_key=st.session_state.settings["gemini_api_key"])
                        model = genai.GenerativeModel('gemini-2.5-flash')
                        full_p = f"Write a professional email for a CRM template. Subject: {t_subject}. Prompt: {ai_prompt}. Use {{name}} for personalization. No subject in body, just message."
                        response = model.generate_content(full_p)
                        st.session_state[f"ai_gen_{edit_id or 'new'}"] = response.text
                    except Exception as e:
                        st.error(f"AI Error: {e}")

            default_body = st.session_state.get(f"ai_gen_{edit_id or 'new'}", initial_body)
            t_body = st.text_area("Message Body (Plain Text or HTML)", value=default_body, placeholder="Aap apni aam zubaan mein yahan likh sakte hain...", height=300)
            
            st.markdown("""
                <div style="display:flex; gap:10px; margin-top:-10px; margin-bottom:15px;">
                    <span style="font-size:0.65rem; color:#1b6656; background:rgba(27,102,86,0.1); padding:2px 8px; border-radius:4px; font-weight:700;">{{name}} = Lead Name</span>
                    <span style="font-size:0.65rem; color:#1b6656; background:rgba(27,102,86,0.1); padding:2px 8px; border-radius:4px; font-weight:700;">{{company}} = Company</span>
                </div>
            """, unsafe_allow_html=True)
            
            use_premium = st.toggle("Apply Premium Enterprise Styling", value=True, help="Aapke plain text ko aik stylish corporate email mein convert kar dega.")
            st.markdown('</div>', unsafe_allow_html=True)
            
            c1, c2, c3 = st.columns([2, 1, 1])
            with c1:
                submit = st.form_submit_button("💾 UNLEASH / SAVE TEMPLATE", use_container_width=True)
            with c2:
                preview = st.form_submit_button("👁️ PREVIEW DESIGN", use_container_width=True)
            with c3:
                if edit_id:
                    if st.form_submit_button("❌ CANCEL", use_container_width=True):
                        st.session_state.edit_template_id = None
                        st.rerun()

            if preview:
                final_html = get_premium_email_layout(t_subject, t_body) if use_premium else t_body
                st.markdown("### 👁️ Visual Preview")
                st.components.v1.html(final_html, height=500, scrolling=True)

            if submit:
                if t_name and t_body:
                    # We save RAW text now, send_email will handle the wrapping
                    payload = {"name": t_name, "subject": t_subject, "body": t_body, "created_at": str(datetime.now())}
                    if edit_id:
                        db.update_template(edit_id, payload)
                        st.toast("✅ Template Synchronized!")
                    else:
                        db.add_template(payload)
                        st.toast("✨ New Protocol Template Authorized!")
                    st.session_state.edit_template_id = None
                    st.session_state.pop(f"ai_gen_{edit_id or 'new'}", None)
                    st.rerun()
                else:
                    st.error("Operational Requirement: Title and Content are mandatory.")

        st.markdown("<br>", unsafe_allow_html=True)

        # --- Template List ---
        if not templates:
            st.info("No templates found in database.")
        else:
            for t in templates:
                with st.expander(f"📄 {t['name']}"):
                    st.markdown(f"**Subject:** {t['subject']}")
                    st.code(t['body'], language="html")
                    
                    # Control Row
                    col1, col2 = st.columns(2)
                    if col1.button("✏️ Edit Template", key=f"e_{t['id']}", use_container_width=True):
                        st.session_state.edit_template_id = t['id']
                        st.session_state.delete_confirm_id = None
                        st.rerun()
                    
                    if col2.button("🗑️ Delete Template", key=f"d_{t['id']}", use_container_width=True):
                        st.session_state.delete_confirm_id = t['id']
                        st.rerun()
                    
                    # Conditional Delete Confirmation (Appears below buttons)
                    if st.session_state.get("delete_confirm_id") == t['id']:
                        st.markdown("---")
                        st.error(f"**Permanently delete '{t['name']}'?**")
                        cc1, cc2 = st.columns(2)
                        if cc1.button("CONFIRM DELETE", key=f"conf_y_{t['id']}", use_container_width=True):
                            db.delete_template(t['id'])
                            st.session_state.delete_confirm_id = None
                            st.success("Template deleted.")
                            st.rerun()
                        if cc2.button("BACK / CANCEL", key=f"conf_n_{t['id']}", use_container_width=True):
                            st.session_state.delete_confirm_id = None
                            st.rerun()

    with tab1:
        st.markdown('<div class="section-heading">📥 High-Volume Bulk Dispatch</div>', unsafe_allow_html=True)
        st.markdown('<div class="luxury-container" style="padding:25px;">', unsafe_allow_html=True)
        
        uploaded_file = st.file_uploader("Upload External Distribution List (Excel/CSV)", type=["xlsx", "csv", "xls"])
        
        if uploaded_file:
            try:
                if uploaded_file.name.endswith('.csv'):
                    df_bulk = pd.read_csv(uploaded_file)
                else:
                    df_bulk = pd.read_excel(uploaded_file)
                
                st.write("📊 **File Preview (First 5 Rows):**")
                st.dataframe(df_bulk.head(), use_container_width=True)
                
                cols = df_bulk.columns.tolist()
                c_email, c_name = st.columns(2)
                email_col = c_email.selectbox("Identify Email Column", cols, index=0)
                name_col = c_name.selectbox("Identify Name Column (Optional)", ["None"] + cols, index=0)
                
                # Compose from Template or Manual
                templates = db.get_all_templates()
                t_opts = ["— Manual Composition —"] + [t['name'] for t in templates]
                sel_bt = st.selectbox("Select Campaign Template", t_opts, key="bt_sel")
                
                b_subj = ""
                b_body = ""
                if sel_bt != "— Manual Composition —":
                    st_match = next(t for t in templates if t['name'] == sel_bt)
                    b_subj = st_match['subject']
                    b_body = st_match['body']
                
                final_b_subj = st.text_input("Broadcast Subject", value=b_subj)
                final_b_body = st.text_area("Broadcast Message (HTML)", value=b_body, height=200)
                
                if st.button("🚀 INITIATE BULK DISPATCH PROTOCOL", use_container_width=True):
                    if not final_b_subj or not final_b_body:
                        st.error("Subject and Body are required.")
                    else:
                        bulk_targets = []
                        for _, row in df_bulk.iterrows():
                            e_val = str(row[email_col]).strip()
                            if "@" in e_val:
                                n_val = str(row[name_col]) if name_col != "None" else "Client"
                                bulk_targets.append({"email": e_val, "name": n_val})
                        
                        if not bulk_targets:
                            st.error("No valid emails found in the selected column.")
                        else:
                            # Log as a mega campaign
                            camp_id = db.add_campaign({
                                "name": f"Bulk_Import_{uploaded_file.name}_{date.today()}",
                                "template_id": None,
                                "subject": final_b_subj,
                                "body": final_b_body,
                                "status": "Running",
                                "created_at": str(datetime.now())
                            })
                            
                            p_bar = st.progress(0, text="Starting Bulk Relay...")
                            ok_count = 0
                            fail_count = 0
                            
                            for i, t in enumerate(bulk_targets):
                                ready_body = final_b_body.replace("{{name}}", t['name'])
                                ok, msg = send_email(final_b_subj, ready_body, t['email'])
                                
                                if ok:
                                    ok_count += 1
                                    db.add_campaign_log(camp_id, t['email'], "Success")
                                else:
                                    fail_count += 1
                                    db.add_campaign_log(camp_id, t['email'], "Failed", msg)
                                
                                p_bar.progress(int(((i + 1) / len(bulk_targets)) * 100), 
                                               text=f"Processing {i+1}/{len(bulk_targets)} -> {t['email']}")
                                time.sleep(1.5) # Anti-spam delay
                            
                            db.update_campaign_stats(camp_id, ok_count, fail_count)
                            p_bar.empty()
                            st.success(f"Bulk Dispatch Complete! Sent: {ok_count}, Failed: {fail_count}")
                            st.balloons()
                            
            except Exception as e:
                st.error(f"Error processing file: {e}")
        else:
            st.info("Upload an Excel/CSV file to begin bulk processing.")
        st.markdown('</div>', unsafe_allow_html=True)

    with tab3:
        st.markdown('<div class="section-heading">📊 Historical Intelligence</div>', unsafe_allow_html=True)
        campaigns = db.get_all_campaigns()
        if not campaigns:
            st.info("No campaign activity detected in history.")
        else:
            for camp in campaigns:
                with st.expander(f"📦 {camp['name']} ({camp['status']})"):
                    st.markdown(f"**Date:** {camp['created_at']}")
                    st.markdown(f"**Subject:** {camp['subject']}")
                    c1, c2, c3 = st.columns(3)
                    c1.metric("🚀 Sent", camp['stats_sent'])
                    c2.metric("❌ Failed", camp['stats_failed'])
                    total = camp['stats_sent'] + camp['stats_failed']
                    c3.metric("📈 Success Rate", f"{(camp['stats_sent']/total*100):.1f}%" if total > 0 else "0%")
                    
                    if st.button("👁️ View Logs", key=f"vlogs_{camp['id']}"):
                        clogs = db.get_campaign_logs(camp['id'])
                        if clogs:
                            st.table(pd.DataFrame(clogs)[["email", "status", "error_message", "sent_at"]])
                        else:
                            st.write("No detailed logs for this campaign.")

elif page == "👤 User Management":
    st.markdown("""
    <div class='page-header animate-in'>
        <div class='page-title'>User Governance Protocol</div>
        <div class='page-sub'>Role-Based Access Control & Force Deployment</div>
    </div>
    """, unsafe_allow_html=True)

    allowed_list = st.session_state.allowed_pages.split(",") if st.session_state.allowed_pages else []
    if st.session_state.role != "Admin" and "👤 User Management" not in allowed_list:
        st.error("Access Denied: You do not have sufficient clearance for User Governance.")
        st.stop()

    if "edit_user_id" not in st.session_state:
        st.session_state.edit_user_id = None

    tab1, tab2 = st.tabs(["➕ Add New Operator", "👥 Active Forces"])

    with tab1:
        st.markdown('<div class="section-heading">✦ Initialize New Personnel</div>', unsafe_allow_html=True)
        with st.form("add_user_form"):
            st.markdown('<div class="luxury-container" style="padding:25px;">', unsafe_allow_html=True)
            u_name = st.text_input("Username / ID", placeholder="e.g., sales_agent_1")
            u_pass = st.text_input("Assigned Password", type="password", placeholder="Enter secure password")
            u_role = st.selectbox("Operational Role", ["User", "Admin"])
            
            st.write("🔓 **Authorized Page Access**")
            p_dash = st.checkbox("📊 Dashboard", value=True)
            p_sales = st.checkbox("💰 Sales Report", value=False)
            p_leads = st.checkbox("👥 Leads", value=True)
            p_tasks = st.checkbox("✅ Tasks", value=True)
            p_reminders = st.checkbox("📧 Reminders", value=False)
            p_marketing = st.checkbox("📢 Email Marketing", value=False)
            p_settings = st.checkbox("⚙️ Settings", value=False)
            p_users = st.checkbox("👤 User Management", value=False)
            
            st.markdown('</div>', unsafe_allow_html=True)
            submit = st.form_submit_button("🚀 DEPLOY PERSONNEL", use_container_width=True)

            if submit:
                if u_name and u_pass:
                    allowed = []
                    if p_dash: allowed.append("📊 Dashboard")
                    if p_sales: allowed.append("💰 Sales Report")
                    if p_leads: allowed.append("👥 Leads")
                    if p_tasks: allowed.append("✅ Tasks")
                    if p_reminders: allowed.append("📧 Reminders")
                    if p_marketing: allowed.append("📢 Email Marketing")
                    if p_settings: allowed.append("⚙️ Settings")
                    if p_users: allowed.append("👤 User Management")
                    
                    payload = {
                        "username": u_name,
                        "password": u_pass,
                        "role": u_role,
                        "allowed_pages": ",".join(allowed),
                        "created_at": str(datetime.now())
                    }
                    if db.add_user(payload):
                        st.success(f"Personnel {u_name} successfully deployed to active duty!")
                        st.rerun()
                    else:
                        st.error("Deployment Failed: Username already exists in protocol.")
                else:
                    st.error("Operational Requirement: All fields must be populated.")

    with tab2:
        st.markdown('<div class="section-heading">✦ Active Personnel Overview</div>', unsafe_allow_html=True)
        users_list = db.get_all_users()
        if not users_list:
            st.info("No active personnel detected.")
        else:
            # --- USER PAGINATION ENGINE ---
            U_PS = 10
            total_up = (len(users_list) // U_PS) + (1 if len(users_list) % U_PS > 0 else 0)
            if 'user_page' not in st.session_state: st.session_state.user_page = 1
            if st.session_state.user_page > total_up: st.session_state.user_page = max(1, total_up)
            
            if total_up > 1:
                uc1, uc2, uc3 = st.columns([1, 2, 1])
                with uc1:
                    if st.button("⬅️ PREVIOUS", key="prev_user", disabled=st.session_state.user_page <= 1, use_container_width=True):
                        st.session_state.user_page -= 1
                        st.rerun()
                with uc2:
                    st.markdown(f"<div style='text-align:center; padding-top:10px; font-weight:800; color:#1b6656; font-size:1rem;'>OPERATOR PAGE {st.session_state.user_page} / {total_up}</div>", unsafe_allow_html=True)
                with uc3:
                    if st.button("NEXT ➡️", key="next_user", disabled=st.session_state.user_page >= total_up, use_container_width=True):
                        st.session_state.user_page += 1
                        st.rerun()
                
                u_start = (st.session_state.user_page - 1) * U_PS
                display_users = users_list[u_start:u_start + U_PS]
            else:
                display_users = users_list

            for usr in display_users:
                with st.expander(f"👤 {usr['username']} - [{usr['role']}]"):
                    st.write(f"**Permissions:** {usr['allowed_pages'] or 'No Pages Authorized'}")
                    st.write(f"**Deployed Since:** {usr['created_at']}")
                    
                    c1, c2 = st.columns(2)
                    if c1.button(f"✏️ Edit Protocol", key=f"edit_u_{usr['id']}", use_container_width=True):
                        st.session_state.edit_user_id = usr['id']
                        st.rerun()

                    if usr['username'] != st.session_state.username:
                        if c2.button(f"🗑️ Terminate", key=f"del_u_{usr['id']}", use_container_width=True):
                            confirm_delete_user(usr['id'], usr['username'])
                    else:
                        c2.info("Self-termination disabled.")

                    # --- Edit Form Section ---
                    if st.session_state.edit_user_id == usr['id']:
                        st.markdown("---")
                        st.markdown(f"#### ⚙️ Re-Configuring: {usr['username']}")
                        with st.form(f"edit_form_{usr['id']}"):
                            is_admin = (usr['role'] == "Admin")
                            
                            new_u = st.text_input("Username", value=usr['username'], disabled=is_admin)
                            new_p = st.text_input("New Password", type="password", placeholder="Enter new secure password")
                            
                            if not is_admin:
                                new_r = st.selectbox("Role", ["User", "Admin"], index=0 if usr['role'] == "User" else 1)
                                st.write("🔓 Update Permissions")
                                allowed_list = usr['allowed_pages'].split(",") if usr['allowed_pages'] else []
                                e_dash = st.checkbox("📊 Dashboard", value="📊 Dashboard" in allowed_list)
                                e_sales = st.checkbox("💰 Sales Report", value="💰 Sales Report" in allowed_list)
                                e_leads = st.checkbox("👥 Leads", value="👥 Leads" in allowed_list)
                                e_tasks = st.checkbox("✅ Tasks", value="✅ Tasks" in allowed_list)
                                e_reminders = st.checkbox("📧 Reminders", value="📧 Reminders" in allowed_list)
                                e_marketing = st.checkbox("📢 Email Marketing", value="📢 Email Marketing" in allowed_list)
                                e_settings = st.checkbox("⚙️ Settings", value="⚙️ Settings" in allowed_list)
                                e_users = st.checkbox("👤 User Management", value="👤 User Management" in allowed_list)
                            else:
                                st.info("🔒 Admin Protocol: Role and Permissions are locked for security.")

                            cc1, cc2 = st.columns(2)
                            if cc1.form_submit_button("💾 SAVE CHANGES", use_container_width=True):
                                # If Admin, keep original role/perms, else use form values
                                save_role = usr['role'] if is_admin else new_r
                                
                                if is_admin:
                                    save_allowed = usr['allowed_pages']
                                else:
                                    e_allowed = []
                                    if e_dash: e_allowed.append("📊 Dashboard")
                                    if e_sales: e_allowed.append("💰 Sales Report")
                                    if e_leads: e_allowed.append("👥 Leads")
                                    if e_tasks: e_allowed.append("✅ Tasks")
                                    if e_reminders: e_allowed.append("📧 Reminders")
                                    if e_marketing: e_allowed.append("📢 Email Marketing")
                                    if e_settings: e_allowed.append("⚙️ Settings")
                                    if e_users: e_allowed.append("👤 User Management")
                                    save_allowed = ",".join(e_allowed)
                                
                                db.update_user(usr['id'], {
                                    "username": new_u,
                                    "password": new_p if new_p else usr['password'], # Only update if not blank
                                    "role": save_role,
                                    "allowed_pages": save_allowed
                                })
                                st.session_state.edit_user_id = None
                                st.success("Protocol Updated.")
                                st.rerun()
                            
                            if cc2.form_submit_button("❌ CANCEL", use_container_width=True):
                                st.session_state.edit_user_id = None
                                st.rerun()

            # --- BOTTOM PAGINATION FOR USERS ---
            if total_up > 1:
                st.markdown("<br>", unsafe_allow_html=True)
                buc1, buc2, buc3 = st.columns([1, 2, 1])
                with buc1:
                    if st.button("⬅️ PREVIOUS", key="prev_user_bottom", disabled=st.session_state.user_page <= 1, use_container_width=True):
                        st.session_state.user_page -= 1
                        st.rerun()
                with buc2:
                    st.markdown(f"<div style='text-align:center; padding-top:10px; font-weight:800; color:#1b6656; font-size:1rem;'>PAGE {st.session_state.user_page} / {total_up}</div>", unsafe_allow_html=True)
                with buc3:
                    if st.button("NEXT ➡️", key="next_user_bottom", disabled=st.session_state.user_page >= total_up, use_container_width=True):
                        st.session_state.user_page += 1
                        st.rerun()
